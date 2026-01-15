from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any, Literal
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from enum import Enum
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    print("=" * 60)
    print("ERROR: MONGO_URL environment variable is not set!")
    print("Please set MONGO_URL to your MongoDB connection string.")
    print("Example: mongodb+srv://user:pass@cluster.mongodb.net/pos_system")
    print("=" * 60)
    # Use a placeholder that will fail gracefully on first DB access
    mongo_url = "mongodb://localhost:27017/pos_system"

client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000)
db = client[os.environ.get('DB_NAME', 'pos_system')]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'pos-system-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="POS & Inventory Management System", version="1.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== ENUMS ====================

class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    ORG_ADMIN = "org_admin"
    STORE_ADMIN = "store_admin"
    CASHIER = "cashier"

class TaxType(str, Enum):
    STANDARD = "standard"
    ZERO_RATED = "zero_rated"
    EXEMPT = "exempt"

class TransactionStatus(str, Enum):
    PENDING = "pending"
    COMPLETED = "completed"
    VOIDED = "voided"
    REFUNDED = "refunded"

class TransactionType(str, Enum):
    SALE = "sale"
    REFUND = "refund"
    VOID = "void"

class PaymentMethod(str, Enum):
    CASH = "cash"
    MOBILE_MONEY = "mobile_money"
    CARD = "card"
    MIXED = "mixed"

class StockMovementType(str, Enum):
    STOCK_IN = "stock_in"
    STOCK_OUT = "stock_out"
    ADJUSTMENT = "adjustment"
    TRANSFER_OUT = "transfer_out"
    TRANSFER_IN = "transfer_in"
    DAMAGE = "damage"
    EXPIRY = "expiry"
    SALE = "sale"
    RETURN = "return"

class TransferStatus(str, Enum):
    DRAFT = "draft"
    DISPATCHED = "dispatched"
    IN_TRANSIT = "in_transit"
    RECEIVED = "received"
    CANCELLED = "cancelled"

# ==================== MODELS ====================

# Base Models
class TimestampMixin(BaseModel):
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Organization Models
class PaymentMethodConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str  # cash, card, mobile_money, custom_*
    is_active: bool = True
    requires_reference: bool = False
    icon: str = "💵"

class PrinterConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str = "thermal"  # thermal, receipt, label
    connection: str = "usb"  # usb, bluetooth, network
    address: str = ""  # IP address or device path
    paper_width: int = 80  # mm
    is_default: bool = False

class OrganizationSettings(BaseModel):
    currency: str = "ZMW"
    currency_symbol: str = "K"
    tax_rate: float = 16.0  # Zambia VAT rate
    tax_inclusive_pricing: bool = True
    tpin: Optional[str] = None
    receipt_footer: str = "Thank you for your business!"
    invoice_prefix: str = ""
    timezone: str = "Africa/Lusaka"
    date_format: str = "DD/MM/YYYY"
    fiscal_compliance_enabled: bool = True
    # Configurable payment methods
    payment_methods: List[PaymentMethodConfig] = Field(default_factory=lambda: [
        PaymentMethodConfig(name="Cash", code="cash", icon="💵", is_active=True),
        PaymentMethodConfig(name="Card", code="card", icon="💳", is_active=True, requires_reference=True),
        PaymentMethodConfig(name="Mobile Money", code="mobile_money", icon="📱", is_active=True, requires_reference=True),
    ])
    # Printer configurations
    printers: List[PrinterConfig] = Field(default_factory=list)
    # Stock settings
    allow_negative_stock: bool = False
    low_stock_threshold: int = 10
    # Logo settings (stored as base64 data URLs)
    system_logo: Optional[str] = None  # Logo displayed throughout the app
    invoice_logo: Optional[str] = None  # Logo displayed on invoices/receipts

class Organization(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    slug: str
    settings: OrganizationSettings = Field(default_factory=OrganizationSettings)
    is_active: bool = True
    subscription_plan: str = "free"  # free, basic, premium, enterprise
    max_stores: int = 1
    max_users: int = 5

class OrganizationCreate(BaseModel):
    name: str
    slug: str
    settings: Optional[OrganizationSettings] = None
    subscription_plan: str = "free"
    max_stores: int = 1
    max_users: int = 5

# Store Address & Location Models (used by both Stores and Warehouses)
class StoreAddress(BaseModel):
    street: str = ""
    city: str = ""
    province: str = ""
    postal_code: str = ""
    country: str = "Zambia"

class StoreLocation(BaseModel):
    latitude: Optional[float] = None
    longitude: Optional[float] = None

# Warehouse Models
class Warehouse(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    name: str
    code: str
    address: StoreAddress = Field(default_factory=StoreAddress)
    location: StoreLocation = Field(default_factory=StoreLocation)
    phone: str = ""
    email: str = ""
    is_active: bool = True
    is_central: bool = False  # Mark as central warehouse
    manager_id: Optional[str] = None

class WarehouseCreate(BaseModel):
    name: str
    code: str
    address: Optional[StoreAddress] = None
    location: Optional[StoreLocation] = None
    phone: str = ""
    email: str = ""
    is_central: bool = False

# Warehouse Stock Model
class WarehouseStock(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    warehouse_id: str
    product_id: str
    quantity: float = 0
    reorder_level: float = 50
    reorder_quantity: float = 100

class WarehouseStockMovement(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    warehouse_id: str
    product_id: str
    movement_type: StockMovementType
    quantity: float
    reference_id: Optional[str] = None
    reason: str = ""
    user_id: str
    balance_after: float = 0

# Warehouse to Store Transfer
class WarehouseTransferItem(BaseModel):
    product_id: str
    product_name: str
    sku: str
    quantity_requested: float
    quantity_dispatched: float = 0
    quantity_received: float = 0

class WarehouseTransfer(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    source_warehouse_id: str
    destination_store_id: Optional[str] = None  # If transferring to store
    destination_warehouse_id: Optional[str] = None  # If transferring to another warehouse
    transfer_number: str
    transfer_type: str = "warehouse_to_store"  # warehouse_to_store, warehouse_to_warehouse
    status: TransferStatus = TransferStatus.DRAFT
    items: List[WarehouseTransferItem] = Field(default_factory=list)
    requested_by: Optional[str] = None
    requested_at: Optional[datetime] = None
    dispatched_by: Optional[str] = None
    dispatched_at: Optional[datetime] = None
    received_by: Optional[str] = None
    received_at: Optional[datetime] = None
    notes: str = ""

class WarehouseTransferCreate(BaseModel):
    source_warehouse_id: str
    destination_store_id: Optional[str] = None
    destination_warehouse_id: Optional[str] = None
    items: List[WarehouseTransferItem]
    notes: str = ""

# Store Models
class Store(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    name: str
    code: str
    address: StoreAddress = Field(default_factory=StoreAddress)
    location: StoreLocation = Field(default_factory=StoreLocation)
    phone: str = ""
    email: str = ""
    is_active: bool = True
    last_sync_at: Optional[datetime] = None

class StoreCreate(BaseModel):
    name: str
    code: str
    address: Optional[StoreAddress] = None
    location: Optional[StoreLocation] = None
    phone: str = ""
    email: str = ""

# User Models
class User(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: Optional[str] = None
    store_ids: List[str] = Field(default_factory=list)
    email: EmailStr
    password_hash: str
    first_name: str
    last_name: str
    role: UserRole
    is_active: bool = True
    last_login: Optional[datetime] = None
    pin: Optional[str] = None  # For quick cashier login

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    first_name: str
    last_name: str
    role: UserRole
    store_ids: List[str] = Field(default_factory=list)
    pin: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    organization_id: Optional[str]
    store_ids: List[str]
    email: str
    first_name: str
    last_name: str
    role: UserRole
    is_active: bool

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

# Product Models
class ProductVariant(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    sku: str
    barcode: Optional[str] = None
    cost_price: float
    selling_price: float
    is_active: bool = True

class Product(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    name: str
    description: str = ""
    sku: str
    barcode: Optional[str] = None
    brand: str = ""  # NEW: Brand field
    category: str = "General"
    cost_price: float
    selling_price: float
    tax_type: TaxType = TaxType.EXEMPT
    unit: str = "piece"
    variants: List[ProductVariant] = Field(default_factory=list)
    is_active: bool = True
    image_base64: Optional[str] = None

class ProductCreate(BaseModel):
    name: str
    description: str = ""
    sku: str
    barcode: Optional[str] = None
    brand: str = ""  # NEW: Brand field
    category: str = "General"
    cost_price: float
    selling_price: float
    tax_type: TaxType = TaxType.EXEMPT
    unit: str = "piece"
    variants: List[ProductVariant] = Field(default_factory=list)
    image_base64: Optional[str] = None

# Store-Specific Pricing with Audit Trail
class StorePricingAudit(BaseModel):
    changed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    changed_by: str
    old_price: Optional[float] = None
    new_price: float
    reason: str = ""

class StorePricing(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    product_id: str
    selling_price: float
    cost_price: Optional[float] = None  # Optional cost override
    is_active: bool = True
    audit_trail: List[StorePricingAudit] = Field(default_factory=list)

class StorePricingCreate(BaseModel):
    product_id: str
    selling_price: float
    cost_price: Optional[float] = None
    reason: str = ""

# Credit Note Models
class CreditNoteItem(BaseModel):
    product_id: str
    product_name: str
    sku: str
    quantity: float
    unit_price: float
    line_total: float

class CreditNote(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    credit_note_number: str
    original_transaction_id: str
    original_receipt_number: str
    items: List[CreditNoteItem] = Field(default_factory=list)
    subtotal: float = 0
    tax_amount: float = 0
    total: float = 0
    reason: str = ""
    status: str = "issued"  # issued, used, cancelled
    issued_by: str
    issued_by_name: str
    used_transaction_id: Optional[str] = None  # If used in a new transaction

class CreditNoteCreate(BaseModel):
    original_transaction_id: str
    items: List[CreditNoteItem]
    reason: str = ""

# Goods Received Note
class GoodsReceivedNote(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    grn_number: str
    transfer_id: str
    transfer_number: str
    source_warehouse_id: str
    source_warehouse_name: str
    items: List[Dict[str, Any]] = Field(default_factory=list)
    received_by: str
    received_by_name: str
    notes: str = ""

# Day Report Models
class DayReportProductSummary(BaseModel):
    product_id: str
    product_name: str
    sku: str
    quantity_sold: float
    total_revenue: float

class DayReportPaymentSummary(BaseModel):
    method: str
    method_name: str
    transaction_count: int
    total_amount: float

class DayReport(BaseModel):
    session_id: str
    store_id: str
    store_name: str
    cashier_name: str
    session_start: str
    session_end: str
    opening_balance: float
    closing_balance: float
    expected_balance: float
    variance: float
    total_sales: float
    total_refunds: float
    transaction_count: int
    products_sold: List[DayReportProductSummary]
    payment_summary: List[DayReportPaymentSummary]

# Stock Models
class Stock(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    product_id: str
    quantity: float = 0
    reorder_level: float = 10
    reorder_quantity: float = 50

class StockMovement(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    product_id: str
    movement_type: StockMovementType
    quantity: float
    reference_id: Optional[str] = None  # Transaction ID, Transfer ID, etc.
    reason: str = ""
    user_id: str
    balance_after: float = 0

class StockMovementCreate(BaseModel):
    product_id: str
    movement_type: StockMovementType
    quantity: float
    reason: str = ""

# Stock Transfer Models
class TransferItem(BaseModel):
    product_id: str
    product_name: str
    sku: str
    quantity_dispatched: float
    quantity_received: float = 0

class StockTransfer(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    source_store_id: str
    destination_store_id: str
    transfer_number: str
    status: TransferStatus = TransferStatus.DRAFT
    items: List[TransferItem] = Field(default_factory=list)
    dispatched_by: Optional[str] = None
    dispatched_at: Optional[datetime] = None
    received_by: Optional[str] = None
    received_at: Optional[datetime] = None
    notes: str = ""

class StockTransferCreate(BaseModel):
    source_store_id: str
    destination_store_id: str
    items: List[TransferItem]
    notes: str = ""

# Transaction Models
class TransactionItem(BaseModel):
    product_id: str
    product_name: str
    sku: str
    brand: Optional[str] = ""
    quantity: float
    unit_price: float
    discount_amount: float = 0
    tax_type: TaxType = TaxType.EXEMPT
    tax_amount: float = 0
    line_total: float = 0

class Payment(BaseModel):
    method: PaymentMethod
    amount: float
    reference: Optional[str] = ""

class Transaction(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    receipt_number: str
    transaction_type: TransactionType = TransactionType.SALE
    status: TransactionStatus = TransactionStatus.COMPLETED
    items: List[TransactionItem] = Field(default_factory=list)
    subtotal: float = 0
    discount_amount: float = 0
    tax_amount: float = 0
    total: float = 0
    payments: List[Payment] = Field(default_factory=list)
    cashier_id: str
    cashier_name: str
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    notes: str = ""
    voided_reason: Optional[str] = None
    refund_reason: Optional[str] = None
    original_transaction_id: Optional[str] = None
    local_id: Optional[str] = None  # For offline sync
    synced: bool = True

class TransactionCreate(BaseModel):
    items: List[TransactionItem]
    payments: List[Payment]
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    notes: str = ""
    discount_amount: float = 0
    local_id: Optional[str] = None

# Cashier Session Models
class CashierSession(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    cashier_id: str
    cashier_name: str
    opening_balance: float = 0
    closing_balance: Optional[float] = None
    expected_balance: Optional[float] = None
    variance: Optional[float] = None
    started_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    ended_at: Optional[datetime] = None
    is_active: bool = True
    total_sales: float = 0
    total_refunds: float = 0

# Print Log Models
class PrintLog(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: Optional[str] = None
    session_id: Optional[str] = None
    user_id: Optional[str] = None
    device_id: Optional[str] = None
    log_level: str  # 'info', 'warn', 'error', 'debug'
    log_tag: str  # e.g., 'AUTO-PRINT', 'PRINT-RECEIPT', 'SEND-RAWBT'
    message: str
    data: Optional[Dict[str, Any]] = None
    receipt_number: Optional[str] = None
    transaction_id: Optional[str] = None
    user_agent: Optional[str] = None
    platform: Optional[str] = None
    error: Optional[Dict[str, Any]] = None

class PrintLogCreate(BaseModel):
    log_level: str
    log_tag: str
    message: str
    data: Optional[Dict[str, Any]] = None
    receipt_number: Optional[str] = None
    transaction_id: Optional[str] = None
    device_id: Optional[str] = None
    session_id: Optional[str] = None
    transaction_count: int = 0

class CashierSessionCreate(BaseModel):
    opening_balance: float = 0

class CashierSessionClose(BaseModel):
    closing_balance: float

# Stock Audit Models
class AuditItem(BaseModel):
    product_id: str
    product_name: str
    sku: str
    system_quantity: float
    counted_quantity: float
    variance: float = 0

class StockAudit(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    audit_number: str
    status: str = "draft"  # draft, in_progress, completed, approved
    items: List[AuditItem] = Field(default_factory=list)
    started_by: str
    started_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_by: Optional[str] = None
    completed_at: Optional[datetime] = None
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    notes: str = ""

# Analytics Models
class DailySalesSummary(BaseModel):
    date: str
    store_id: str
    store_name: str
    gross_sales: float = 0
    net_sales: float = 0
    tax_collected: float = 0
    transaction_count: int = 0
    average_transaction: float = 0
    refund_count: int = 0
    refund_amount: float = 0

class StoreSalesAnalytics(BaseModel):
    store_id: str
    store_name: str
    location: Optional[StoreLocation] = None
    daily: float = 0
    weekly: float = 0
    monthly: float = 0
    yearly: float = 0
    transaction_count: int = 0

# Sync Models
class SyncLog(TimestampMixin):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    store_id: str
    sync_type: str  # push, pull
    entity_type: str  # transactions, products, stock
    records_synced: int = 0
    status: str = "pending"  # pending, in_progress, completed, failed
    error_message: Optional[str] = None

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, organization_id: Optional[str], role: str) -> str:
    payload = {
        "sub": user_id,
        "org": organization_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> Dict[str, Any]:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Dict[str, Any]:
    try:
        payload = decode_token(credentials.credentials)
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            logger.warning(f"User not found for token payload: {payload.get('sub')}")
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        logger.warning("Token expired")
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError as e:
        logger.warning(f"Invalid token: {str(e)}")
        raise HTTPException(status_code=401, detail="Invalid token")
    except Exception as e:
        logger.error(f"Error getting current user: {str(e)}", exc_info=True)
        raise HTTPException(status_code=401, detail="Authentication failed")

def require_role(allowed_roles: List[UserRole]):
    async def role_checker(user: Dict = Depends(get_current_user)):
        if UserRole(user["role"]) not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

def serialize_datetime(obj):
    """Convert datetime objects to ISO strings for MongoDB storage"""
    if isinstance(obj, dict):
        return {k: serialize_datetime(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [serialize_datetime(item) for item in obj]
    elif isinstance(obj, datetime):
        return obj.isoformat()
    return obj

def deserialize_datetime(obj, fields=['created_at', 'updated_at', 'started_at', 'ended_at', 'dispatched_at', 'received_at', 'last_login', 'last_sync_at', 'completed_at', 'approved_at']):
    """Convert ISO strings back to datetime objects"""
    if isinstance(obj, dict):
        # Create a copy to avoid modifying the original
        result = {}
        for key, value in obj.items():
            if key in fields and isinstance(value, str):
                try:
                    # Handle different datetime formats
                    if 'T' in value or ' ' in value:
                        # Try parsing ISO format
                        if value.endswith('Z'):
                            value = value[:-1] + '+00:00'
                        result[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        # Keep original value if not a valid datetime string
                        result[key] = value
                except (ValueError, AttributeError, TypeError):
                    # Keep original value if parsing fails
                    result[key] = value
            elif isinstance(value, dict):
                result[key] = deserialize_datetime(value, fields)
            elif isinstance(value, list):
                result[key] = [deserialize_datetime(item, fields) for item in value]
            else:
                result[key] = value
        return result
    elif isinstance(obj, list):
        return [deserialize_datetime(item, fields) for item in obj]
    return obj

async def generate_receipt_number(org_id: str, store_id: str) -> str:
    """Generate sequential receipt number"""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0, "settings": 1})
    prefix_raw = ""
    if org:
        prefix_raw = org.get("settings", {}).get("invoice_prefix") or ""
    prefix = prefix_raw.strip().replace(" ", "")
    if prefix:
        prefix = prefix.rstrip("-")
        prefix = f"{prefix}-"
    store = await db.stores.find_one({"id": store_id}, {"_id": 0, "code": 1})
    store_code = ""
    if store:
        store_code = (store.get("code") or "").strip().replace(" ", "")
    store_segment = f"{store_code}-" if store_code else ""
    prefix_with_date = f"{prefix}{store_segment}{today}"
    count = await db.transactions.count_documents({
        "organization_id": org_id,
        "store_id": store_id,
        "receipt_number": {"$regex": f"^{re.escape(prefix_with_date)}"}
    })
    return f"{prefix_with_date}-{str(count + 1).zfill(6)}"

async def generate_transfer_number(org_id: str) -> str:
    """Generate sequential transfer number"""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.stock_transfers.count_documents({
        "organization_id": org_id,
        "transfer_number": {"$regex": f"^TR-{today}"}
    })
    return f"TR-{today}-{str(count + 1).zfill(4)}"

async def generate_audit_number(org_id: str) -> str:
    """Generate sequential audit number"""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.stock_audits.count_documents({
        "organization_id": org_id,
        "audit_number": {"$regex": f"^AUD-{today}"}
    })
    return f"AUD-{today}-{str(count + 1).zfill(4)}"

def calculate_tax(amount: float, tax_type: TaxType, tax_rate: float, inclusive: bool = True) -> Dict[str, float]:
    """Calculate tax based on type and settings"""
    if tax_type == TaxType.EXEMPT or tax_type == TaxType.ZERO_RATED:
        return {"tax_amount": 0, "net_amount": amount}
    
    if inclusive:
        net_amount = amount / (1 + tax_rate / 100)
        tax_amount = amount - net_amount
    else:
        tax_amount = amount * (tax_rate / 100)
        net_amount = amount
    
    return {"tax_amount": round(tax_amount, 2), "net_amount": round(net_amount, 2)}

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserCreate):
    """Register a new user (creates organization for org_admin)"""
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        email=data.email,
        password_hash=hash_password(data.password),
        first_name=data.first_name,
        last_name=data.last_name,
        role=data.role,
        store_ids=data.store_ids,
        pin=data.pin
    )
    
    # If registering as org_admin, create a default organization
    if data.role == UserRole.ORG_ADMIN:
        org = Organization(
            name=f"{data.first_name}'s Organization",
            slug=data.email.split('@')[0].lower().replace('.', '-')
        )
        await db.organizations.insert_one(serialize_datetime(org.model_dump()))
        user.organization_id = org.id
    
    await db.users.insert_one(serialize_datetime(user.model_dump()))
    
    token = create_token(user.id, user.organization_id, user.role)
    return TokenResponse(
        access_token=token,
        user=UserResponse(**user.model_dump())
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    """Login with email and password"""
    logger.info(f"Login attempt for email: {data.email}")
    try:
        user = await db.users.find_one({"email": data.email}, {"_id": 0})
        if not user:
            logger.warning(f"Login failed: User not found for email {data.email}")
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        if not verify_password(data.password, user["password_hash"]):
            logger.warning(f"Login failed: Invalid password for email {data.email}")
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        if not user.get("is_active", True):
            logger.warning(f"Login failed: Account deactivated for email {data.email}")
            raise HTTPException(status_code=401, detail="Account is deactivated")
        
        # Update last login
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
        )
        
        token = create_token(user["id"], user.get("organization_id"), user["role"])
        logger.info(f"Login successful for user: {user['id']} ({data.email})")
        return TokenResponse(
            access_token=token,
            user=UserResponse(**user)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: Dict = Depends(get_current_user)):
    """Get current user profile"""
    return UserResponse(**user)

# ==================== ORGANIZATION ROUTES ====================

@api_router.post("/organizations", response_model=Organization)
async def create_organization(
    data: OrganizationCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Create a new organization (Super Admin only)"""
    existing = await db.organizations.find_one({"slug": data.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Organization slug already exists")
    
    # Filter out None values to let Organization model use defaults
    org_data = {k: v for k, v in data.model_dump().items() if v is not None}
    org = Organization(**org_data)
    await db.organizations.insert_one(serialize_datetime(org.model_dump()))
    return org

@api_router.get("/organizations", response_model=List[Organization])
async def get_organizations(user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))):
    """Get all organizations (Super Admin only)"""
    orgs = await db.organizations.find({}, {"_id": 0}).to_list(1000)
    return [deserialize_datetime(org) for org in orgs]

@api_router.get("/organizations/current", response_model=Organization)
async def get_current_organization(user: Dict = Depends(get_current_user)):
    """Get current user's organization"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=404, detail="No organization associated")
    
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    return deserialize_datetime(org)

@api_router.put("/organizations/{org_id}", response_model=Organization)
async def update_organization(
    org_id: str,
    data: OrganizationCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Update organization"""
    if user["role"] == UserRole.ORG_ADMIN and user.get("organization_id") != org_id:
        raise HTTPException(status_code=403, detail="Cannot update other organizations")
    
    update_data = data.model_dump(exclude_unset=True)
    if "settings" in update_data and isinstance(update_data.get("settings"), dict):
        existing = await db.organizations.find_one({"id": org_id}, {"_id": 0, "settings": 1})
        if existing:
            update_data["settings"] = {
                **existing.get("settings", {}),
                **update_data.get("settings", {})
            }
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.organizations.update_one(
        {"id": org_id},
        {"$set": serialize_datetime(update_data)}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    return deserialize_datetime(org)

# ==================== STORE ROUTES ====================

@api_router.post("/stores", response_model=Store)
async def create_store(
    data: StoreCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Create a new store"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    existing = await db.stores.find_one({
        "organization_id": user["organization_id"],
        "code": data.code
    })
    if existing:
        raise HTTPException(status_code=400, detail="Store code already exists")
    
    # Prepare store data with defaults for optional fields
    store_data = data.model_dump()
    if store_data.get("address") is None:
        store_data["address"] = StoreAddress()
    if store_data.get("location") is None:
        store_data["location"] = StoreLocation()
    
    store = Store(
        organization_id=user["organization_id"],
        **store_data
    )
    await db.stores.insert_one(serialize_datetime(store.model_dump()))
    return store

@api_router.get("/stores", response_model=List[Store])
async def get_stores(user: Dict = Depends(get_current_user)):
    """Get all stores for current organization"""
    try:
        logger.info(f"Getting stores for user: {user.get('id')}, org: {user.get('organization_id')}, role: {user.get('role')}")
        
        if not user.get("organization_id"):
            logger.warning(f"User {user.get('id')} has no organization_id")
            raise HTTPException(status_code=400, detail="No organization associated. Please contact your administrator.")
        
        query = {"organization_id": user["organization_id"]}
        
        # Cashiers and store admins only see their assigned stores
        if user["role"] in [UserRole.CASHIER, UserRole.STORE_ADMIN]:
            if user.get("store_ids"):
                query["id"] = {"$in": user["store_ids"]}
            else:
                # No assigned stores, return empty list
                logger.info(f"User {user.get('id')} has no assigned stores")
                return []
        
        logger.info(f"Querying stores with query: {query}")
        stores = await db.stores.find(query, {"_id": 0}).to_list(1000)
        logger.info(f"Found {len(stores)} stores for organization {user.get('organization_id')}")
        
        if not stores:
            logger.info("No stores found, returning empty list")
            return []
        
        # Deserialize stores and convert to Store models
        result_stores = []
        for idx, store in enumerate(stores):
            try:
                # Ensure required fields exist with proper defaults
                if 'address' not in store or store['address'] is None:
                    store['address'] = StoreAddress().model_dump()
                elif isinstance(store['address'], dict):
                    # Ensure address dict has all required fields
                    address = store['address']
                    store['address'] = {
                        'street': address.get('street', ''),
                        'city': address.get('city', ''),
                        'province': address.get('province', ''),
                        'postal_code': address.get('postal_code', ''),
                        'country': address.get('country', 'Zambia')
                    }
                
                if 'location' not in store or store['location'] is None:
                    store['location'] = StoreLocation().model_dump()
                elif isinstance(store['location'], dict):
                    # Ensure location dict has all required fields
                    location = store['location']
                    store['location'] = {
                        'latitude': location.get('latitude'),
                        'longitude': location.get('longitude')
                    }
                
                # Deserialize datetime fields
                deserialized = deserialize_datetime(store)
                
                # Ensure datetime fields are either datetime objects or None
                for field in ['created_at', 'updated_at', 'last_sync_at']:
                    if field in deserialized:
                        if deserialized[field] is None:
                            # Use current time for None timestamps
                            if field in ['created_at', 'updated_at']:
                                deserialized[field] = datetime.now(timezone.utc)
                        elif isinstance(deserialized[field], str):
                            try:
                                deserialized[field] = datetime.fromisoformat(deserialized[field].replace('Z', '+00:00'))
                            except:
                                deserialized[field] = datetime.now(timezone.utc) if field in ['created_at', 'updated_at'] else None
                
                # Convert to Store model instance
                store_model = Store(**deserialized)
                result_stores.append(store_model)
            except Exception as e:
                logger.error(f"Error processing store {idx}: {store.get('id', 'unknown')}: {str(e)}", exc_info=True)
                # Skip problematic stores rather than failing entirely
                continue
        
        logger.info(f"Returning {len(result_stores)} stores")
        return result_stores
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_stores: {str(e)}", exc_info=True)
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@api_router.get("/stores/{store_id}", response_model=Store)
async def get_store(store_id: str, user: Dict = Depends(get_current_user)):
    """Get store by ID"""
    query = {"id": store_id}
    # Organization admins can only access stores from their organization
    if user["role"] != UserRole.SUPER_ADMIN:
        if not user.get("organization_id"):
            raise HTTPException(status_code=400, detail="No organization associated")
        query["organization_id"] = user["organization_id"]
    
    store = await db.stores.find_one(query, {"_id": 0})
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    return deserialize_datetime(store)

@api_router.put("/stores/{store_id}", response_model=Store)
async def update_store(
    store_id: str,
    data: StoreCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Update store"""
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.stores.update_one(
        {"id": store_id, "organization_id": user["organization_id"]},
        {"$set": serialize_datetime(update_data)}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Store not found")
    
    store = await db.stores.find_one({"id": store_id}, {"_id": 0})
    return deserialize_datetime(store)

# ==================== WAREHOUSE ROUTES ====================

@api_router.post("/warehouses", response_model=Warehouse)
async def create_warehouse(
    data: WarehouseCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Create a new warehouse"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    existing = await db.warehouses.find_one({
        "organization_id": user["organization_id"],
        "code": data.code
    })
    if existing:
        raise HTTPException(status_code=400, detail="Warehouse code already exists")
    
    # Prepare warehouse data with defaults for optional fields
    warehouse_data = data.model_dump()
    if warehouse_data.get("address") is None:
        warehouse_data["address"] = StoreAddress()
    if warehouse_data.get("location") is None:
        warehouse_data["location"] = StoreLocation()
    
    warehouse = Warehouse(
        organization_id=user["organization_id"],
        **warehouse_data
    )
    await db.warehouses.insert_one(serialize_datetime(warehouse.model_dump()))
    return warehouse

@api_router.get("/warehouses", response_model=List[Warehouse])
async def get_warehouses(user: Dict = Depends(get_current_user)):
    """Get all warehouses for current organization"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    warehouses = await db.warehouses.find(
        {"organization_id": user["organization_id"]}, 
        {"_id": 0}
    ).to_list(1000)
    return [deserialize_datetime(w) for w in warehouses]

@api_router.get("/warehouses/{warehouse_id}", response_model=Warehouse)
async def get_warehouse(warehouse_id: str, user: Dict = Depends(get_current_user)):
    """Get warehouse by ID"""
    warehouse = await db.warehouses.find_one({
        "id": warehouse_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not warehouse:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    return deserialize_datetime(warehouse)

@api_router.put("/warehouses/{warehouse_id}", response_model=Warehouse)
async def update_warehouse(
    warehouse_id: str,
    data: WarehouseCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Update warehouse"""
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.warehouses.update_one(
        {"id": warehouse_id, "organization_id": user["organization_id"]},
        {"$set": serialize_datetime(update_data)}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    return deserialize_datetime(warehouse)

@api_router.delete("/warehouses/{warehouse_id}")
async def delete_warehouse(
    warehouse_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Deactivate warehouse"""
    result = await db.warehouses.update_one(
        {"id": warehouse_id, "organization_id": user["organization_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    return {"message": "Warehouse deactivated"}

# Warehouse Stock Management
@api_router.get("/warehouses/{warehouse_id}/stock")
async def get_warehouse_stock(warehouse_id: str, user: Dict = Depends(get_current_user)):
    """Get stock levels for a warehouse"""
    pipeline = [
        {"$match": {"warehouse_id": warehouse_id, "organization_id": user["organization_id"]}},
        {"$lookup": {
            "from": "products",
            "localField": "product_id",
            "foreignField": "id",
            "as": "product"
        }},
        {"$unwind": "$product"},
        {"$project": {
            "_id": 0,
            "id": 1,
            "product_id": 1,
            "quantity": 1,
            "reorder_level": 1,
            "reorder_quantity": 1,
            "product_name": "$product.name",
            "sku": "$product.sku",
            "barcode": "$product.barcode",
            "cost_price": "$product.cost_price",
            "selling_price": "$product.selling_price"
        }}
    ]
    
    stock = await db.warehouse_stock.aggregate(pipeline).to_list(1000)
    return stock

@api_router.post("/warehouses/{warehouse_id}/stock/movement")
async def create_warehouse_stock_movement(
    warehouse_id: str,
    data: StockMovementCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Create stock movement for warehouse"""
    # Get current stock
    stock = await db.warehouse_stock.find_one({
        "warehouse_id": warehouse_id,
        "product_id": data.product_id,
        "organization_id": user["organization_id"]
    })
    
    current_qty = stock["quantity"] if stock else 0
    
    # Calculate new quantity
    if data.movement_type in [StockMovementType.STOCK_IN, StockMovementType.TRANSFER_IN, StockMovementType.RETURN]:
        new_qty = current_qty + data.quantity
    elif data.movement_type in [StockMovementType.STOCK_OUT, StockMovementType.TRANSFER_OUT, StockMovementType.DAMAGE, StockMovementType.EXPIRY]:
        new_qty = current_qty - data.quantity
        if new_qty < 0:
            raise HTTPException(status_code=400, detail="Insufficient warehouse stock")
    else:  # ADJUSTMENT
        new_qty = data.quantity
    
    # Update or create stock record
    if stock:
        await db.warehouse_stock.update_one(
            {"id": stock["id"]},
            {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        new_stock = WarehouseStock(
            organization_id=user["organization_id"],
            warehouse_id=warehouse_id,
            product_id=data.product_id,
            quantity=new_qty
        )
        await db.warehouse_stock.insert_one(serialize_datetime(new_stock.model_dump()))
    
    # Create movement record
    movement = WarehouseStockMovement(
        organization_id=user["organization_id"],
        warehouse_id=warehouse_id,
        product_id=data.product_id,
        movement_type=data.movement_type,
        quantity=data.quantity,
        reason=data.reason,
        user_id=user["id"],
        balance_after=new_qty
    )
    await db.warehouse_stock_movements.insert_one(serialize_datetime(movement.model_dump()))
    
    return {"success": True, "new_quantity": new_qty}

@api_router.get("/warehouses/{warehouse_id}/stock/movements")
async def get_warehouse_stock_movements(
    warehouse_id: str,
    product_id: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Get stock movements for a warehouse"""
    query = {"warehouse_id": warehouse_id, "organization_id": user["organization_id"]}
    if product_id:
        query["product_id"] = product_id
    
    movements = await db.warehouse_stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [deserialize_datetime(m) for m in movements]

# Warehouse Transfers
async def generate_warehouse_transfer_number(org_id: str) -> str:
    """Generate sequential warehouse transfer number"""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.warehouse_transfers.count_documents({
        "organization_id": org_id,
        "transfer_number": {"$regex": f"^WT-{today}"}
    })
    return f"WT-{today}-{str(count + 1).zfill(4)}"

@api_router.post("/warehouse-transfers", response_model=WarehouseTransfer)
async def create_warehouse_transfer(
    data: WarehouseTransferCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Create warehouse transfer (to store or another warehouse)"""
    transfer_number = await generate_warehouse_transfer_number(user["organization_id"])
    
    transfer_type = "warehouse_to_store" if data.destination_store_id else "warehouse_to_warehouse"
    
    transfer = WarehouseTransfer(
        organization_id=user["organization_id"],
        source_warehouse_id=data.source_warehouse_id,
        destination_store_id=data.destination_store_id,
        destination_warehouse_id=data.destination_warehouse_id,
        transfer_number=transfer_number,
        transfer_type=transfer_type,
        items=data.items,
        notes=data.notes,
        requested_by=user["id"],
        requested_at=datetime.now(timezone.utc)
    )
    await db.warehouse_transfers.insert_one(serialize_datetime(transfer.model_dump()))
    return transfer

@api_router.get("/warehouse-transfers")
async def get_warehouse_transfers(
    warehouse_id: Optional[str] = None,
    store_id: Optional[str] = None,
    status: Optional[TransferStatus] = None,
    user: Dict = Depends(get_current_user)
):
    """Get warehouse transfers"""
    query = {"organization_id": user["organization_id"]}
    
    if warehouse_id:
        query["$or"] = [
            {"source_warehouse_id": warehouse_id}, 
            {"destination_warehouse_id": warehouse_id}
        ]
    
    if store_id:
        query["destination_store_id"] = store_id
    
    if status:
        query["status"] = status.value
    
    transfers = await db.warehouse_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [deserialize_datetime(t) for t in transfers]

@api_router.put("/warehouse-transfers/{transfer_id}/dispatch")
async def dispatch_warehouse_transfer(
    transfer_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Dispatch a warehouse transfer"""
    transfer = await db.warehouse_transfers.find_one({
        "id": transfer_id,
        "organization_id": user["organization_id"]
    })
    
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    if transfer["status"] != TransferStatus.DRAFT.value:
        raise HTTPException(status_code=400, detail="Can only dispatch draft transfers")
    
    # Deduct stock from source warehouse
    for item in transfer["items"]:
        qty_to_dispatch = item.get("quantity_requested", item.get("quantity_dispatched", 0))
        
        # Check warehouse stock
        wh_stock = await db.warehouse_stock.find_one({
            "warehouse_id": transfer["source_warehouse_id"],
            "product_id": item["product_id"],
            "organization_id": user["organization_id"]
        })
        
        if not wh_stock or wh_stock["quantity"] < qty_to_dispatch:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient warehouse stock for {item['product_name']}"
            )
        
        # Deduct from warehouse
        await db.warehouse_stock.update_one(
            {"id": wh_stock["id"]},
            {"$inc": {"quantity": -qty_to_dispatch}}
        )
        
        # Record movement
        movement = WarehouseStockMovement(
            organization_id=user["organization_id"],
            warehouse_id=transfer["source_warehouse_id"],
            product_id=item["product_id"],
            movement_type=StockMovementType.TRANSFER_OUT,
            quantity=qty_to_dispatch,
            reference_id=transfer_id,
            reason=f"Transfer {transfer['transfer_number']}",
            user_id=user["id"],
            balance_after=wh_stock["quantity"] - qty_to_dispatch
        )
        await db.warehouse_stock_movements.insert_one(serialize_datetime(movement.model_dump()))
        
        item["quantity_dispatched"] = qty_to_dispatch
    
    # Update transfer
    await db.warehouse_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": TransferStatus.DISPATCHED.value,
            "items": transfer["items"],
            "dispatched_by": user["id"],
            "dispatched_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.warehouse_transfers.find_one({"id": transfer_id}, {"_id": 0})
    return deserialize_datetime(updated)

@api_router.put("/warehouse-transfers/{transfer_id}/receive")
async def receive_warehouse_transfer(
    transfer_id: str,
    received_items: List[Dict[str, Any]],
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Receive a warehouse transfer"""
    transfer = await db.warehouse_transfers.find_one({
        "id": transfer_id,
        "organization_id": user["organization_id"]
    })
    
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    if transfer["status"] != TransferStatus.DISPATCHED.value:
        raise HTTPException(status_code=400, detail="Can only receive dispatched transfers")
    
    # Add stock to destination
    for received in received_items:
        for item in transfer["items"]:
            if item["product_id"] == received["product_id"]:
                item["quantity_received"] = received["quantity_received"]
                
                if transfer["transfer_type"] == "warehouse_to_store":
                    # Add to store stock
                    movement_data = StockMovementCreate(
                        product_id=item["product_id"],
                        movement_type=StockMovementType.TRANSFER_IN,
                        quantity=received["quantity_received"],
                        reason=f"Warehouse Transfer {transfer['transfer_number']}"
                    )
                    await create_stock_movement(transfer["destination_store_id"], movement_data, user)
                else:
                    # Add to destination warehouse
                    dest_stock = await db.warehouse_stock.find_one({
                        "warehouse_id": transfer["destination_warehouse_id"],
                        "product_id": item["product_id"],
                        "organization_id": user["organization_id"]
                    })
                    
                    if dest_stock:
                        await db.warehouse_stock.update_one(
                            {"id": dest_stock["id"]},
                            {"$inc": {"quantity": received["quantity_received"]}}
                        )
                        new_qty = dest_stock["quantity"] + received["quantity_received"]
                    else:
                        new_stock = WarehouseStock(
                            organization_id=user["organization_id"],
                            warehouse_id=transfer["destination_warehouse_id"],
                            product_id=item["product_id"],
                            quantity=received["quantity_received"]
                        )
                        await db.warehouse_stock.insert_one(serialize_datetime(new_stock.model_dump()))
                        new_qty = received["quantity_received"]
                    
                    # Record movement
                    movement = WarehouseStockMovement(
                        organization_id=user["organization_id"],
                        warehouse_id=transfer["destination_warehouse_id"],
                        product_id=item["product_id"],
                        movement_type=StockMovementType.TRANSFER_IN,
                        quantity=received["quantity_received"],
                        reference_id=transfer_id,
                        reason=f"Transfer {transfer['transfer_number']}",
                        user_id=user["id"],
                        balance_after=new_qty
                    )
                    await db.warehouse_stock_movements.insert_one(serialize_datetime(movement.model_dump()))
    
    await db.warehouse_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": TransferStatus.RECEIVED.value,
            "items": transfer["items"],
            "received_by": user["id"],
            "received_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.warehouse_transfers.find_one({"id": transfer_id}, {"_id": 0})
    return deserialize_datetime(updated)

# ==================== USER MANAGEMENT ROUTES ====================

@api_router.post("/users", response_model=UserResponse)
async def create_user(
    data: UserCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Create a new user"""
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate role hierarchy
    creator_role = UserRole(user["role"])
    target_role = data.role
    
    if creator_role == UserRole.STORE_ADMIN and target_role not in [UserRole.CASHIER]:
        raise HTTPException(status_code=403, detail="Store admins can only create cashiers")
    
    if creator_role == UserRole.ORG_ADMIN and target_role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Cannot create super admin")
    
    new_user = User(
        organization_id=user.get("organization_id"),
        email=data.email,
        password_hash=hash_password(data.password),
        first_name=data.first_name,
        last_name=data.last_name,
        role=target_role,
        store_ids=data.store_ids,
        pin=data.pin
    )
    
    await db.users.insert_one(serialize_datetime(new_user.model_dump()))
    return UserResponse(**new_user.model_dump())

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get all users for current organization"""
    query = {}
    if user["role"] != UserRole.SUPER_ADMIN:
        query["organization_id"] = user.get("organization_id")
    
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [UserResponse(**deserialize_datetime(u)) for u in users]

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: str,
    data: Dict[str, Any],
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Update user"""
    if "password" in data:
        data["password_hash"] = hash_password(data.pop("password"))
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    query = {"id": user_id}
    if user["role"] != UserRole.SUPER_ADMIN:
        query["organization_id"] = user.get("organization_id")
    
    result = await db.users.update_one(query, {"$set": serialize_datetime(data)})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return UserResponse(**deserialize_datetime(updated_user))

# ==================== PRODUCT ROUTES ====================

@api_router.post("/products", response_model=Product)
async def create_product(
    data: ProductCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Create a new product"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    existing = await db.products.find_one({
        "organization_id": user["organization_id"],
        "$or": [{"sku": data.sku}, {"barcode": data.barcode}] if data.barcode else [{"sku": data.sku}]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Product with same SKU or barcode exists")
    
    product = Product(
        organization_id=user["organization_id"],
        **data.model_dump()
    )
    await db.products.insert_one(serialize_datetime(product.model_dump()))
    return product

@api_router.get("/products", response_model=List[Product])
async def get_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Get all products for current organization"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    query = {"organization_id": user["organization_id"], "is_active": True}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"barcode": search}
        ]
    
    if category:
        query["category"] = category
    
    products = await db.products.find(query, {"_id": 0}).to_list(1000)
    return [deserialize_datetime(p) for p in products]

@api_router.get("/products/brands")
async def get_product_brands(user: Dict = Depends(get_current_user)):
    """Get all unique brands for filtering"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    brands = await db.products.distinct("brand", {
        "organization_id": user["organization_id"],
        "is_active": True,
        "brand": {"$ne": ""}
    })
    return sorted([b for b in brands if b])  # Filter out empty strings and None

@api_router.get("/products/template")
async def download_product_template(user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))):
    """Download Excel template for product import"""
    logger.info("Generating product import template")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Header row with styling
        headers = [
            "Name*", "SKU*", "Barcode", "Brand", "Category", 
            "Cost Price*", "Selling Price*", "Tax Type", "Unit", "Description"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add example rows
        examples = [
            ["Sample Product 1", "SKU001", "1234567890123", "Brand A", "Electronics", "100.00", "150.00", "exempt", "piece", "Sample description"],
            ["Sample Product 2", "SKU002", "", "Brand B", "Food", "50.00", "75.00", "standard", "kg", ""],
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Set column widths
        column_widths = [25, 15, 18, 15, 15, 12, 15, 12, 10, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Product Import Template - Instructions"],
            [""],
            ["Required Fields (marked with *):"],
            ["- Name: Product name (required)"],
            ["- SKU: Stock Keeping Unit, must be unique (required)"],
            ["- Cost Price: Product cost price (required)"],
            ["- Selling Price: Product selling price (required)"],
            [""],
            ["Optional Fields:"],
            ["- Barcode: Product barcode"],
            ["- Brand: Product brand name"],
            ["- Category: Product category (default: General)"],
            ["- Tax Type: exempt (default), zero_rated, or standard"],
            ["- Unit: piece (default), kg, g, l, ml, box, pack"],
            ["- Description: Product description"],
            [""],
            ["Tax Type Options:"],
            ["- exempt: No tax (default)"],
            ["- zero_rated: 0% tax"],
            ["- standard: VAT 16%"],
            [""],
            ["Unit Options:"],
            ["- piece, kg, g, l, ml, box, pack"],
            [""],
            ["Notes:"],
            ["- SKU must be unique within your organization"],
            ["- Barcode must be unique if provided"],
            ["- Prices should be numeric values"],
            ["- Remove example rows before uploading"],
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction[0])
            if row_num == 1:
                ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Get bytes from the buffer
        excel_bytes = output.getvalue()
        output.close()
        
        logger.info(f"Template generated successfully, size: {len(excel_bytes)} bytes")
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=product_import_template.xlsx",
                "Content-Length": str(len(excel_bytes))
            }
        )
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {str(e)}")

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, user: Dict = Depends(get_current_user)):
    """Get product by ID"""
    product = await db.products.find_one({
        "id": product_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return deserialize_datetime(product)

@api_router.get("/products/barcode/{barcode}", response_model=Product)
async def get_product_by_barcode(barcode: str, user: Dict = Depends(get_current_user)):
    """Get product by barcode"""
    product = await db.products.find_one({
        "barcode": barcode,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return deserialize_datetime(product)

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(
    product_id: str,
    data: ProductCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Update product"""
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.products.update_one(
        {"id": product_id, "organization_id": user["organization_id"]},
        {"$set": serialize_datetime(update_data)}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Verify product belongs to user's organization
    query = {"id": product_id}
    if user["role"] != UserRole.SUPER_ADMIN:
        query["organization_id"] = user["organization_id"]
    product = await db.products.find_one(query, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return deserialize_datetime(product)

@api_router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Soft delete product"""
    result = await db.products.update_one(
        {"id": product_id, "organization_id": user["organization_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted"}

@api_router.post("/products/import")
async def import_products_from_excel(
    file: UploadFile = File(...),
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Import products from Excel file"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        
        # Validate required columns
        required_columns = ['Name', 'SKU', 'Cost Price', 'Selling Price']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Process products
        created_count = 0
        updated_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Get values with defaults
                name = str(row['Name']).strip()
                sku = str(row['SKU']).strip()
                barcode = str(row.get('Barcode', '')).strip() if pd.notna(row.get('Barcode')) else None
                brand = str(row.get('Brand', '')).strip() if pd.notna(row.get('Brand')) else ''
                category = str(row.get('Category', 'General')).strip() if pd.notna(row.get('Category')) else 'General'
                cost_price = float(row['Cost Price'])
                selling_price = float(row['Selling Price'])
                tax_type = str(row.get('Tax Type', 'exempt')).strip().lower() if pd.notna(row.get('Tax Type')) else 'exempt'
                unit = str(row.get('Unit', 'piece')).strip().lower() if pd.notna(row.get('Unit')) else 'piece'
                description = str(row.get('Description', '')).strip() if pd.notna(row.get('Description')) else ''
                
                # Validate tax_type
                if tax_type not in ['exempt', 'zero_rated', 'standard']:
                    tax_type = 'exempt'
                
                # Validate unit
                valid_units = ['piece', 'kg', 'g', 'l', 'ml', 'box', 'pack']
                if unit not in valid_units:
                    unit = 'piece'
                
                # Validate required fields
                if not name or not sku:
                    errors.append(f"Row {index + 2}: Name and SKU are required")
                    continue
                
                if cost_price < 0 or selling_price < 0:
                    errors.append(f"Row {index + 2}: Prices must be non-negative")
                    continue
                
                # Check if product exists
                query = {
                    "organization_id": user["organization_id"],
                    "sku": sku
                }
                if barcode:
                    query = {
                        "organization_id": user["organization_id"],
                        "$or": [{"sku": sku}, {"barcode": barcode}]
                    }
                
                existing = await db.products.find_one(query)
                
                product_data = {
                    "name": name,
                    "description": description,
                    "sku": sku,
                    "barcode": barcode if barcode else None,
                    "brand": brand,
                    "category": category,
                    "cost_price": cost_price,
                    "selling_price": selling_price,
                    "tax_type": tax_type,
                    "unit": unit,
                    "is_active": True,
                }
                
                if existing:
                    # Update existing product
                    update_data = serialize_datetime({
                        **product_data,
                        "updated_at": datetime.now(timezone.utc)
                    })
                    await db.products.update_one(
                        {"id": existing["id"]},
                        {"$set": update_data}
                    )
                    updated_count += 1
                else:
                    # Create new product
                    product = Product(
                        organization_id=user["organization_id"],
                        **product_data
                    )
                    await db.products.insert_one(serialize_datetime(product.model_dump()))
                    created_count += 1
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue
        
        return {
            "message": "Import completed",
            "created": created_count,
            "updated": updated_count,
            "errors": errors,
            "total_processed": created_count + updated_count
        }
        
    except Exception as e:
        logger.error(f"Error importing products: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.get("/products/with-stock/{store_id}")
async def get_products_with_stock(
    store_id: str,
    search: Optional[str] = None,
    category: Optional[str] = None,
    brand: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Get all products with their stock levels and store-specific pricing for POS"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    match_query = {"organization_id": user["organization_id"], "is_active": True}
    
    if search:
        match_query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"barcode": search},
            {"brand": {"$regex": search, "$options": "i"}}
        ]
    
    if category:
        match_query["category"] = category
    
    if brand:
        match_query["brand"] = {"$regex": brand, "$options": "i"}
    
    pipeline = [
        {"$match": match_query},
        {"$lookup": {
            "from": "stock",
            "let": {"product_id": "$id"},
            "pipeline": [
                {"$match": {
                    "$expr": {
                        "$and": [
                            {"$eq": ["$product_id", "$$product_id"]},
                            {"$eq": ["$store_id", store_id]}
                        ]
                    }
                }}
            ],
            "as": "stock_info"
        }},
        {"$lookup": {
            "from": "store_pricing",
            "let": {"product_id": "$id"},
            "pipeline": [
                {"$match": {
                    "$expr": {
                        "$and": [
                            {"$eq": ["$product_id", "$$product_id"]},
                            {"$eq": ["$store_id", store_id]},
                            {"$eq": ["$is_active", True]}
                        ]
                    }
                }}
            ],
            "as": "store_pricing"
        }},
        {"$addFields": {
            "stock_quantity": {
                "$ifNull": [{"$arrayElemAt": ["$stock_info.quantity", 0]}, 0]
            },
            "reorder_level": {
                "$ifNull": [{"$arrayElemAt": ["$stock_info.reorder_level", 0]}, 10]
            },
            "store_selling_price": {
                "$ifNull": [{"$arrayElemAt": ["$store_pricing.selling_price", 0]}, "$selling_price"]
            },
            "has_store_price": {
                "$gt": [{"$size": "$store_pricing"}, 0]
            }
        }},
        {"$project": {"stock_info": 0, "store_pricing": 0, "_id": 0}}
    ]
    
    products = await db.products.aggregate(pipeline).to_list(1000)
    return products

# ==================== STORE PRICING ROUTES ====================

@api_router.get("/stores/{store_id}/pricing")
async def get_store_pricing(
    store_id: str,
    user: Dict = Depends(get_current_user)
):
    """Get all store-specific pricing"""
    pipeline = [
        {"$match": {"store_id": store_id, "organization_id": user["organization_id"], "is_active": True}},
        {"$lookup": {
            "from": "products",
            "localField": "product_id",
            "foreignField": "id",
            "as": "product"
        }},
        {"$unwind": "$product"},
        {"$project": {
            "_id": 0,
            "id": 1,
            "product_id": 1,
            "selling_price": 1,
            "cost_price": 1,
            "audit_trail": 1,
            "product_name": "$product.name",
            "sku": "$product.sku",
            "default_price": "$product.selling_price",
            "brand": "$product.brand"
        }}
    ]
    
    pricing = await db.store_pricing.aggregate(pipeline).to_list(1000)
    return pricing

@api_router.post("/stores/{store_id}/pricing")
async def set_store_pricing(
    store_id: str,
    data: StorePricingCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Set store-specific pricing for a product"""
    # Check if pricing already exists
    existing = await db.store_pricing.find_one({
        "store_id": store_id,
        "product_id": data.product_id,
        "organization_id": user["organization_id"]
    })
    
    audit_entry = StorePricingAudit(
        changed_by=user["id"],
        old_price=existing["selling_price"] if existing else None,
        new_price=data.selling_price,
        reason=data.reason
    )
    
    if existing:
        # Update existing pricing
        audit_trail = existing.get("audit_trail", [])
        audit_trail.append(serialize_datetime(audit_entry.model_dump()))
        
        await db.store_pricing.update_one(
            {"id": existing["id"]},
            {"$set": {
                "selling_price": data.selling_price,
                "cost_price": data.cost_price,
                "audit_trail": audit_trail,
                "is_active": True,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated = await db.store_pricing.find_one({"id": existing["id"]}, {"_id": 0})
        return updated
    else:
        # Create new pricing
        pricing = StorePricing(
            organization_id=user["organization_id"],
            store_id=store_id,
            product_id=data.product_id,
            selling_price=data.selling_price,
            cost_price=data.cost_price,
            audit_trail=[audit_entry]
        )
        await db.store_pricing.insert_one(serialize_datetime(pricing.model_dump()))
        return pricing

@api_router.delete("/stores/{store_id}/pricing/{product_id}")
async def remove_store_pricing(
    store_id: str,
    product_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Remove store-specific pricing (revert to default)"""
    result = await db.store_pricing.update_one(
        {"store_id": store_id, "product_id": product_id, "organization_id": user["organization_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pricing not found")
    
    return {"message": "Store pricing removed"}

@api_router.get("/stores/{store_id}/pricing/{product_id}/audit")
async def get_pricing_audit(
    store_id: str,
    product_id: str,
    user: Dict = Depends(get_current_user)
):
    """Get pricing audit trail for a product"""
    pricing = await db.store_pricing.find_one({
        "store_id": store_id,
        "product_id": product_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not pricing:
        return {"audit_trail": []}
    
    return {"audit_trail": pricing.get("audit_trail", [])}

# ==================== STOCK ROUTES ====================

@api_router.get("/stock/upload-template")
async def download_stock_upload_template(user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))):
    """Download Excel template for stock upload (Super Admin and Org Admin only)"""
    logger.info("Generating stock upload template")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Upload"
        
        # Header row with styling
        headers = [
            "Store Code*", "SKU*", "Quantity*", "Reorder Level"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add example rows
        examples = [
            ["MAIN001", "SKU001", "100.00", "10.00"],
            ["MAIN001", "SKU002", "50.00", "5.00"],
            ["STORE002", "SKU001", "75.00", "10.00"],
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Set column widths
        column_widths = [15, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Stock Upload Template - Instructions"],
            [""],
            ["Required Fields (marked with *):"],
            ["- Store Code: The code of the store (e.g., MAIN001) (required)"],
            ["- SKU: Stock Keeping Unit of the product (required)"],
            ["- Quantity: Stock quantity to set for the product (required)"],
            [""],
            ["Optional Fields:"],
            ["- Reorder Level: Reorder level for the product (default: 10)"],
            [""],
            ["Notes:"],
            ["- Store Code must match an existing store in your organization"],
            ["- SKU must match an existing product in your organization"],
            ["- Quantity must be a non-negative number"],
            ["- Reorder Level must be a non-negative number"],
            ["- If stock doesn't exist, it will be created"],
            ["- If stock exists, it will be updated"],
            ["- Remove example rows before uploading"],
            [""],
            ["Example:"],
            ["Store Code | SKU    | Quantity | Reorder Level"],
            ["MAIN001   | SKU001 | 100.00   | 10.00"],
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction[0])
            if row_num == 1:
                ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Get bytes from the buffer
        excel_bytes = output.getvalue()
        output.close()
        
        logger.info(f"Stock upload template generated successfully, size: {len(excel_bytes)} bytes")
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=stock_upload_template.xlsx",
                "Content-Length": str(len(excel_bytes))
            }
        )
    except Exception as e:
        logger.error(f"Error generating stock upload template: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate template: {str(e)}")

@api_router.get("/stock/{store_id}", response_model=List[Dict])
async def get_store_stock(store_id: str, user: Dict = Depends(get_current_user)):
    """Get stock levels for a store"""
    # Join stock with products
    pipeline = [
        {"$match": {"store_id": store_id, "organization_id": user["organization_id"]}},
        {"$lookup": {
            "from": "products",
            "localField": "product_id",
            "foreignField": "id",
            "as": "product"
        }},
        {"$unwind": "$product"},
        {"$project": {
            "_id": 0,
            "id": 1,
            "product_id": 1,
            "quantity": 1,
            "reorder_level": 1,
            "product_name": "$product.name",
            "sku": "$product.sku",
            "barcode": "$product.barcode",
            "selling_price": "$product.selling_price"
        }}
    ]
    
    stock = await db.stock.aggregate(pipeline).to_list(1000)
    return stock

@api_router.post("/stock/{store_id}/movement", response_model=StockMovement)
async def create_stock_movement(
    store_id: str,
    data: StockMovementCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Create stock movement (adjust stock)"""
    # Get current stock
    stock = await db.stock.find_one({
        "store_id": store_id,
        "product_id": data.product_id,
        "organization_id": user["organization_id"]
    })
    
    current_qty = stock["quantity"] if stock else 0
    
    # Calculate new quantity
    if data.movement_type in [StockMovementType.STOCK_IN, StockMovementType.TRANSFER_IN, StockMovementType.RETURN]:
        new_qty = current_qty + data.quantity
    elif data.movement_type in [StockMovementType.STOCK_OUT, StockMovementType.TRANSFER_OUT, StockMovementType.SALE, StockMovementType.DAMAGE, StockMovementType.EXPIRY]:
        new_qty = current_qty - data.quantity
        if new_qty < 0:
            raise HTTPException(status_code=400, detail="Insufficient stock")
    else:  # ADJUSTMENT
        new_qty = data.quantity
    
    # Update or create stock record
    if stock:
        await db.stock.update_one(
            {"id": stock["id"]},
            {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        new_stock = Stock(
            organization_id=user["organization_id"],
            store_id=store_id,
            product_id=data.product_id,
            quantity=new_qty
        )
        await db.stock.insert_one(serialize_datetime(new_stock.model_dump()))
    
    # Create movement record
    movement = StockMovement(
        organization_id=user["organization_id"],
        store_id=store_id,
        product_id=data.product_id,
        movement_type=data.movement_type,
        quantity=data.quantity,
        reason=data.reason,
        user_id=user["id"],
        balance_after=new_qty
    )
    await db.stock_movements.insert_one(serialize_datetime(movement.model_dump()))
    
    return movement

@api_router.get("/stock/{store_id}/movements", response_model=List[StockMovement])
async def get_stock_movements(
    store_id: str,
    product_id: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Get stock movements for a store"""
    query = {"store_id": store_id, "organization_id": user["organization_id"]}
    if product_id:
        query["product_id"] = product_id
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [deserialize_datetime(m) for m in movements]

# Store Stock Response Model
class StoreStockItem(BaseModel):
    id: str
    product_id: str
    product_name: str
    sku: str
    barcode: Optional[str] = None
    quantity: float
    reorder_level: float
    selling_price: float

class StoreStockResponse(BaseModel):
    store_id: str
    store_name: str
    store_code: str
    stocks: List[StoreStockItem]

@api_router.get("/stock/all-stores", response_model=List[StoreStockResponse])
async def get_all_stores_stock(user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))):
    """Get stock levels for all stores in the organization (Super Admin and Org Admin only)"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    # Get all stores for the organization
    stores = await db.stores.find(
        {"organization_id": user["organization_id"]}, 
        {"_id": 0}
    ).to_list(1000)
    
    if not stores:
        return []
    
    result = []
    
    # For each store, get its stocks
    for store in stores:
        # Join stock with products for this store
        pipeline = [
            {"$match": {"store_id": store["id"], "organization_id": user["organization_id"]}},
            {"$lookup": {
                "from": "products",
                "localField": "product_id",
                "foreignField": "id",
                "as": "product"
            }},
            {"$unwind": "$product"},
            {"$project": {
                "_id": 0,
                "id": 1,
                "product_id": 1,
                "quantity": 1,
                "reorder_level": 1,
                "product_name": "$product.name",
                "sku": "$product.sku",
                "barcode": "$product.barcode",
                "selling_price": "$product.selling_price"
            }}
        ]
        
        stocks = await db.stock.aggregate(pipeline).to_list(1000)
        
        # Convert to StoreStockItem models
        stock_items = [
            StoreStockItem(
                id=stock["id"],
                product_id=stock["product_id"],
                product_name=stock["product_name"],
                sku=stock["sku"],
                barcode=stock.get("barcode"),
                quantity=stock["quantity"],
                reorder_level=stock["reorder_level"],
                selling_price=stock["selling_price"]
            )
            for stock in stocks
        ]
        
        result.append(StoreStockResponse(
            store_id=store["id"],
            store_name=store["name"],
            store_code=store["code"],
            stocks=stock_items
        ))
    
    return result

@api_router.post("/stock/upload")
async def upload_stocks_from_excel(
    file: UploadFile = File(...),
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Upload stocks for stores from Excel file (Super Admin and Org Admin only)"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        
        # Validate required columns
        required_columns = ['Store Code', 'SKU', 'Quantity']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Get all stores for the organization (for validation)
        stores = await db.stores.find(
            {"organization_id": user["organization_id"]}, 
            {"_id": 0, "id": 1, "code": 1, "name": 1}
        ).to_list(1000)
        
        store_code_map = {store["code"]: store for store in stores}
        
        # Process stocks
        created_count = 0
        updated_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Get values with validation
                store_code = str(row['Store Code']).strip() if pd.notna(row['Store Code']) else None
                sku = str(row['SKU']).strip() if pd.notna(row['SKU']) else None
                quantity = row['Quantity']
                reorder_level = float(row['Reorder Level']) if pd.notna(row.get('Reorder Level')) else 10.0
                
                # Validate required fields
                if not store_code:
                    errors.append(f"Row {index + 2}: Store Code is required")
                    continue
                
                if not sku:
                    errors.append(f"Row {index + 2}: SKU is required")
                    continue
                
                # Validate store exists
                if store_code not in store_code_map:
                    errors.append(f"Row {index + 2}: Store Code '{store_code}' not found in your organization")
                    continue
                
                store = store_code_map[store_code]
                store_id = store["id"]
                
                # Validate quantity
                try:
                    quantity = float(quantity)
                    if quantity < 0:
                        errors.append(f"Row {index + 2}: Quantity must be non-negative")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {index + 2}: Invalid quantity value")
                    continue
                
                # Validate reorder level
                if reorder_level < 0:
                    errors.append(f"Row {index + 2}: Reorder Level must be non-negative")
                    continue
                
                # Find product by SKU
                product = await db.products.find_one({
                    "organization_id": user["organization_id"],
                    "sku": sku
                }, {"_id": 0})
                
                if not product:
                    errors.append(f"Row {index + 2}: Product with SKU '{sku}' not found")
                    continue
                
                product_id = product["id"]
                
                # Check if stock exists
                existing_stock = await db.stock.find_one({
                    "store_id": store_id,
                    "product_id": product_id,
                    "organization_id": user["organization_id"]
                })
                
                if existing_stock:
                    # Update existing stock
                    await db.stock.update_one(
                        {"id": existing_stock["id"]},
                        {"$set": {
                            "quantity": quantity,
                            "reorder_level": reorder_level,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    updated_count += 1
                else:
                    # Create new stock
                    new_stock = Stock(
                        organization_id=user["organization_id"],
                        store_id=store_id,
                        product_id=product_id,
                        quantity=quantity,
                        reorder_level=reorder_level
                    )
                    await db.stock.insert_one(serialize_datetime(new_stock.model_dump()))
                    created_count += 1
                
                # Create stock movement record for tracking
                movement = StockMovement(
                    organization_id=user["organization_id"],
                    store_id=store_id,
                    product_id=product_id,
                    movement_type=StockMovementType.ADJUSTMENT,
                    quantity=quantity,
                    reason=f"Bulk upload from Excel - Row {index + 2}",
                    user_id=user["id"],
                    balance_after=quantity
                )
                await db.stock_movements.insert_one(serialize_datetime(movement.model_dump()))
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue
        
        return {
            "message": "Stock upload completed",
            "created": created_count,
            "updated": updated_count,
            "errors": errors,
            "total_processed": created_count + updated_count,
            "total_rows": len(df)
        }
        
    except Exception as e:
        logger.error(f"Error uploading stocks: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

# ==================== STOCK TRANSFER ROUTES ====================

@api_router.post("/transfers", response_model=StockTransfer)
async def create_transfer(
    data: StockTransferCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Create stock transfer"""
    transfer_number = await generate_transfer_number(user["organization_id"])
    
    transfer = StockTransfer(
        organization_id=user["organization_id"],
        source_store_id=data.source_store_id,
        destination_store_id=data.destination_store_id,
        transfer_number=transfer_number,
        items=data.items,
        notes=data.notes
    )
    await db.stock_transfers.insert_one(serialize_datetime(transfer.model_dump()))
    return transfer

@api_router.get("/transfers", response_model=List[StockTransfer])
async def get_transfers(
    store_id: Optional[str] = None,
    status: Optional[TransferStatus] = None,
    user: Dict = Depends(get_current_user)
):
    """Get stock transfers"""
    query = {"organization_id": user["organization_id"]}
    
    if store_id:
        query["$or"] = [{"source_store_id": store_id}, {"destination_store_id": store_id}]
    
    if status:
        query["status"] = status.value
    
    transfers = await db.stock_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [deserialize_datetime(t) for t in transfers]

@api_router.put("/transfers/{transfer_id}/dispatch", response_model=StockTransfer)
async def dispatch_transfer(
    transfer_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Dispatch a transfer"""
    transfer = await db.stock_transfers.find_one({
        "id": transfer_id,
        "organization_id": user["organization_id"]
    })
    
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    if transfer["status"] != TransferStatus.DRAFT.value:
        raise HTTPException(status_code=400, detail="Can only dispatch draft transfers")
    
    # Deduct stock from source store
    for item in transfer["items"]:
        movement = StockMovementCreate(
            product_id=item["product_id"],
            movement_type=StockMovementType.TRANSFER_OUT,
            quantity=item["quantity_dispatched"],
            reason=f"Transfer {transfer['transfer_number']}"
        )
        await create_stock_movement(transfer["source_store_id"], movement, user)
    
    # Update transfer
    await db.stock_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": TransferStatus.DISPATCHED.value,
            "dispatched_by": user["id"],
            "dispatched_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.stock_transfers.find_one({"id": transfer_id}, {"_id": 0})
    return deserialize_datetime(updated)

@api_router.put("/transfers/{transfer_id}/receive", response_model=StockTransfer)
async def receive_transfer(
    transfer_id: str,
    received_items: List[Dict[str, Any]],
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Receive a transfer (supports partial receiving)"""
    transfer = await db.stock_transfers.find_one({
        "id": transfer_id,
        "organization_id": user["organization_id"]
    })
    
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    if transfer["status"] not in [TransferStatus.DISPATCHED.value, TransferStatus.IN_TRANSIT.value]:
        raise HTTPException(status_code=400, detail="Can only receive dispatched transfers")
    
    # Update received quantities and add stock
    for received in received_items:
        for item in transfer["items"]:
            if item["product_id"] == received["product_id"]:
                item["quantity_received"] = received["quantity_received"]
                
                # Add stock to destination store
                movement = StockMovementCreate(
                    product_id=item["product_id"],
                    movement_type=StockMovementType.TRANSFER_IN,
                    quantity=received["quantity_received"],
                    reason=f"Transfer {transfer['transfer_number']}"
                )
                await create_stock_movement(transfer["destination_store_id"], movement, user)
    
    await db.stock_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": TransferStatus.RECEIVED.value,
            "items": transfer["items"],
            "received_by": user["id"],
            "received_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.stock_transfers.find_one({"id": transfer_id}, {"_id": 0})
    return deserialize_datetime(updated)

# ==================== TRANSACTION (POS) ROUTES ====================

@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(
    data: TransactionCreate,
    store_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """Create a new sale transaction"""
    if not user.get("organization_id"):
        raise HTTPException(status_code=400, detail="No organization associated")
    
    # Get organization settings for tax calculation and stock settings
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    tax_rate = org["settings"]["tax_rate"] if org else 16.0
    tax_inclusive = org["settings"]["tax_inclusive_pricing"] if org else True
    allow_negative_stock = org["settings"].get("allow_negative_stock", False) if org else False
    
    # Validate stock availability for each item
    stock_errors = []
    for item in data.items:
        stock = await db.stock.find_one({
            "store_id": store_id,
            "product_id": item.product_id,
            "organization_id": user["organization_id"]
        })
        current_qty = stock["quantity"] if stock else 0
        
        if not allow_negative_stock and current_qty < item.quantity:
            product = await db.products.find_one({"id": item.product_id}, {"name": 1})
            product_name = product["name"] if product else item.product_id
            stock_errors.append(f"{product_name}: only {current_qty} available, requested {item.quantity}")
    
    if stock_errors:
        raise HTTPException(
            status_code=400, 
            detail={"message": "Insufficient stock", "errors": stock_errors}
        )
    
    # Calculate totals
    subtotal = 0
    total_tax = 0
    
    for item in data.items:
        line_subtotal = item.quantity * item.unit_price - item.discount_amount
        tax_calc = calculate_tax(line_subtotal, item.tax_type, tax_rate, tax_inclusive)
        item.tax_amount = tax_calc["tax_amount"]
        item.line_total = line_subtotal
        subtotal += line_subtotal
        total_tax += item.tax_amount
    
    receipt_number = await generate_receipt_number(user["organization_id"], store_id)
    
    transaction = Transaction(
        organization_id=user["organization_id"],
        store_id=store_id,
        receipt_number=receipt_number,
        items=data.items,
        subtotal=round(subtotal, 2),
        discount_amount=data.discount_amount,
        tax_amount=round(total_tax, 2),
        total=round(subtotal - data.discount_amount, 2),
        payments=data.payments,
        cashier_id=user["id"],
        cashier_name=f"{user['first_name']} {user['last_name']}",
        customer_name=data.customer_name,
        customer_phone=data.customer_phone,
        notes=data.notes,
        local_id=data.local_id
    )
    
    await db.transactions.insert_one(serialize_datetime(transaction.model_dump()))
    
    # Deduct stock for each item
    for item in data.items:
        try:
            movement = StockMovementCreate(
                product_id=item.product_id,
                movement_type=StockMovementType.SALE,
                quantity=item.quantity,
                reason=f"Sale {receipt_number}"
            )
            await create_stock_movement(store_id, movement, user)
        except HTTPException:
            pass  # Continue even if stock update fails (offline scenario)
    
    # Update cashier session if active
    await db.cashier_sessions.update_one(
        {"cashier_id": user["id"], "store_id": store_id, "is_active": True},
        {"$inc": {"total_sales": transaction.total, "transaction_count": 1}}
    )
    
    return transaction

@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(
    store_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[TransactionStatus] = None,
    limit: int = 100,
    user: Dict = Depends(get_current_user)
):
    """Get transactions"""
    query = {"organization_id": user["organization_id"]}
    
    if store_id:
        query["store_id"] = store_id
    
    if status:
        query["status"] = status.value
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return [deserialize_datetime(t) for t in transactions]

@api_router.get("/transactions/{transaction_id}", response_model=Transaction)
async def get_transaction(transaction_id: str, user: Dict = Depends(get_current_user)):
    """Get transaction by ID"""
    transaction = await db.transactions.find_one({
        "id": transaction_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return deserialize_datetime(transaction)

@api_router.post("/transactions/{transaction_id}/void", response_model=Transaction)
async def void_transaction(
    transaction_id: str,
    reason: str = Query(...),
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Void a transaction"""
    transaction = await db.transactions.find_one({
        "id": transaction_id,
        "organization_id": user["organization_id"]
    })
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if transaction["status"] != TransactionStatus.COMPLETED.value:
        raise HTTPException(status_code=400, detail="Can only void completed transactions")
    
    # Restore stock
    for item in transaction["items"]:
        movement = StockMovementCreate(
            product_id=item["product_id"],
            movement_type=StockMovementType.RETURN,
            quantity=item["quantity"],
            reason=f"Void {transaction['receipt_number']}"
        )
        await create_stock_movement(transaction["store_id"], movement, user)
    
    await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": {
            "status": TransactionStatus.VOIDED.value,
            "voided_reason": reason,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    return deserialize_datetime(updated)

@api_router.post("/transactions/{transaction_id}/refund", response_model=Transaction)
async def refund_transaction(
    transaction_id: str,
    reason: str = Query(...),
    items: Optional[List[Dict]] = None,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Refund a transaction (full or partial)"""
    original = await db.transactions.find_one({
        "id": transaction_id,
        "organization_id": user["organization_id"]
    })
    
    if not original:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if original["status"] != TransactionStatus.COMPLETED.value:
        raise HTTPException(status_code=400, detail="Can only refund completed transactions")
    
    # Calculate refund amount
    refund_items = items if items else original["items"]
    refund_total = sum(item.get("line_total", 0) for item in refund_items)
    
    receipt_number = await generate_receipt_number(user["organization_id"], original["store_id"])
    
    # Create refund transaction
    refund = Transaction(
        organization_id=user["organization_id"],
        store_id=original["store_id"],
        receipt_number=receipt_number,
        transaction_type=TransactionType.REFUND,
        items=[TransactionItem(**item) for item in refund_items],
        subtotal=-refund_total,
        total=-refund_total,
        payments=[Payment(method=PaymentMethod.CASH, amount=-refund_total)],
        cashier_id=user["id"],
        cashier_name=f"{user['first_name']} {user['last_name']}",
        refund_reason=reason,
        original_transaction_id=transaction_id
    )
    
    await db.transactions.insert_one(serialize_datetime(refund.model_dump()))
    
    # Restore stock
    for item in refund_items:
        movement = StockMovementCreate(
            product_id=item["product_id"],
            movement_type=StockMovementType.RETURN,
            quantity=item["quantity"],
            reason=f"Refund {receipt_number}"
        )
        await create_stock_movement(original["store_id"], movement, user)
    
    # Mark original as refunded if full refund
    if not items or len(items) == len(original["items"]):
        await db.transactions.update_one(
            {"id": transaction_id},
            {"$set": {"status": TransactionStatus.REFUNDED.value}}
        )
    
    return refund

# ==================== CASHIER SESSION ROUTES ====================

@api_router.post("/sessions/start", response_model=CashierSession)
async def start_cashier_session(
    data: CashierSessionCreate,
    store_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """Start a cashier session"""
    # Check for existing active session
    existing = await db.cashier_sessions.find_one({
        "cashier_id": user["id"],
        "store_id": store_id,
        "is_active": True
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Active session already exists")
    
    session = CashierSession(
        organization_id=user["organization_id"],
        store_id=store_id,
        cashier_id=user["id"],
        cashier_name=f"{user['first_name']} {user['last_name']}",
        opening_balance=data.opening_balance
    )
    
    await db.cashier_sessions.insert_one(serialize_datetime(session.model_dump()))
    return session

@api_router.post("/sessions/end", response_model=CashierSession)
async def end_cashier_session(
    data: CashierSessionClose,
    store_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """End a cashier session"""
    session = await db.cashier_sessions.find_one({
        "cashier_id": user["id"],
        "store_id": store_id,
        "is_active": True
    })
    
    if not session:
        raise HTTPException(status_code=404, detail="No active session found")
    
    # Calculate expected balance
    expected = session["opening_balance"] + session.get("total_sales", 0) - session.get("total_refunds", 0)
    variance = data.closing_balance - expected
    
    await db.cashier_sessions.update_one(
        {"id": session["id"]},
        {"$set": {
            "closing_balance": data.closing_balance,
            "expected_balance": expected,
            "variance": variance,
            "ended_at": datetime.now(timezone.utc).isoformat(),
            "is_active": False,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.cashier_sessions.find_one({"id": session["id"]}, {"_id": 0})
    return deserialize_datetime(updated)

@api_router.get("/sessions/current", response_model=Optional[CashierSession])
async def get_current_session(
    store_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """Get current active session"""
    session = await db.cashier_sessions.find_one({
        "cashier_id": user["id"],
        "store_id": store_id,
        "is_active": True
    }, {"_id": 0})
    
    return deserialize_datetime(session) if session else None

@api_router.get("/sessions", response_model=List[CashierSession])
async def get_sessions(
    store_id: Optional[str] = None,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get cashier sessions"""
    query = {"organization_id": user["organization_id"]}
    if store_id:
        query["store_id"] = store_id
    
    sessions = await db.cashier_sessions.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return [deserialize_datetime(s) for s in sessions]

# ==================== DAY REPORT ROUTES ====================

# ==================== PRINT LOG ROUTES ====================

@api_router.post("/print-logs", response_model=PrintLog)
async def create_print_log(
    data: PrintLogCreate,
    user: Dict = Depends(get_current_user)
):
    """Create a print log entry"""
    # Get current session if available
    session_id = data.session_id
    if not session_id:
        # Try to find active session
        active_session = await db.cashier_sessions.find_one({
            "cashier_id": user["id"],
            "is_active": True
        }, {"id": 1})
        if active_session:
            session_id = active_session["id"]
    
    # Generate device ID from user agent if not provided
    device_id = data.device_id
    if not device_id:
        # Create a simple device fingerprint
        import hashlib
        user_agent = data.data.get("user_agent", "") if data.data else ""
        device_id = hashlib.md5(f"{user_agent}_{user['id']}".encode()).hexdigest()[:16]
    
    print_log = PrintLog(
        organization_id=user["organization_id"],
        store_id=user.get("store_ids", [None])[0] if user.get("store_ids") else None,
        session_id=session_id,
        user_id=user["id"],
        device_id=device_id,
        log_level=data.log_level,
        log_tag=data.log_tag,
        message=data.message,
        data=data.data,
        receipt_number=data.receipt_number,
        transaction_id=data.transaction_id,
        user_agent=data.data.get("user_agent") if data.data else None,
        platform=data.data.get("platform") if data.data else None,
        error=data.data.get("error") if data.data else None
    )
    
    await db.print_logs.insert_one(serialize_datetime(print_log.model_dump()))
    return print_log

@api_router.get("/print-logs", response_model=List[PrintLog])
async def get_print_logs(
    store_id: Optional[str] = None,
    session_id: Optional[str] = None,
    receipt_number: Optional[str] = None,
    log_level: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = Query(100, le=1000),
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get print logs with filters"""
    query = {"organization_id": user["organization_id"]}
    
    if store_id:
        query["store_id"] = store_id
    if session_id:
        query["session_id"] = session_id
    if receipt_number:
        query["receipt_number"] = receipt_number
    if log_level:
        query["log_level"] = log_level
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    logs = await db.print_logs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return [deserialize_datetime(log) for log in logs]

@api_router.get("/sessions/{session_id}/report", response_model=DayReport)
async def get_day_report(
    session_id: str,
    user: Dict = Depends(get_current_user)
):
    """Get detailed day report for a session"""
    session = await db.cashier_sessions.find_one({
        "id": session_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get store info
    store = await db.stores.find_one({"id": session["store_id"]}, {"_id": 0})
    
    # Get all transactions for this session
    session_start = session.get("started_at") or session.get("created_at")
    session_end = session.get("ended_at")
    
    tx_query = {
        "organization_id": user["organization_id"],
        "store_id": session["store_id"],
        "cashier_id": session["cashier_id"],
        "transaction_type": "sale",
        "status": "completed"
    }
    
    # Filter by time range if available
    if session_start:
        tx_query["created_at"] = {"$gte": session_start}
    if session_end:
        if "created_at" in tx_query:
            tx_query["created_at"]["$lte"] = session_end
        else:
            tx_query["created_at"] = {"$lte": session_end}
    
    transactions = await db.transactions.find(tx_query, {"_id": 0}).to_list(1000)
    
    # Calculate products sold summary
    products_sold_dict = {}
    for tx in transactions:
        for item in tx.get("items", []):
            pid = item["product_id"]
            if pid not in products_sold_dict:
                products_sold_dict[pid] = {
                    "product_id": pid,
                    "product_name": item["product_name"],
                    "sku": item["sku"],
                    "quantity_sold": 0,
                    "total_revenue": 0
                }
            products_sold_dict[pid]["quantity_sold"] += item["quantity"]
            products_sold_dict[pid]["total_revenue"] += item["line_total"]
    
    products_sold = list(products_sold_dict.values())
    products_sold.sort(key=lambda x: x["total_revenue"], reverse=True)
    
    # Calculate payment summary
    payment_dict = {}
    payment_names = {"cash": "Cash", "card": "Card", "mobile_money": "Mobile Money"}
    for tx in transactions:
        for payment in tx.get("payments", []):
            method = payment["method"]
            if method not in payment_dict:
                payment_dict[method] = {
                    "method": method,
                    "method_name": payment_names.get(method, method.replace("_", " ").title()),
                    "transaction_count": 0,
                    "total_amount": 0
                }
            payment_dict[method]["transaction_count"] += 1
            payment_dict[method]["total_amount"] += payment["amount"]
    
    payment_summary = list(payment_dict.values())
    payment_summary.sort(key=lambda x: x["total_amount"], reverse=True)
    
    return DayReport(
        session_id=session_id,
        store_id=session["store_id"],
        store_name=store["name"] if store else "Unknown",
        cashier_name=session.get("cashier_name", "Unknown"),
        session_start=session_start or "",
        session_end=session_end or "",
        opening_balance=session.get("opening_balance", 0),
        closing_balance=session.get("closing_balance", 0),
        expected_balance=session.get("expected_balance", 0),
        variance=session.get("variance", 0),
        total_sales=session.get("total_sales", 0),
        total_refunds=session.get("total_refunds", 0),
        transaction_count=len(transactions),
        products_sold=products_sold,
        payment_summary=payment_summary
    )

# ==================== CREDIT NOTE ROUTES ====================

async def generate_credit_note_number(org_id: str) -> str:
    """Generate sequential credit note number"""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.credit_notes.count_documents({
        "organization_id": org_id,
        "credit_note_number": {"$regex": f"^CN-{today}"}
    })
    return f"CN-{today}-{str(count + 1).zfill(4)}"

@api_router.post("/credit-notes", response_model=CreditNote)
async def create_credit_note(
    data: CreditNoteCreate,
    store_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """Create a credit note against an existing transaction (partial return)"""
    # Get original transaction
    original_tx = await db.transactions.find_one({
        "id": data.original_transaction_id,
        "organization_id": user["organization_id"]
    })
    
    if not original_tx:
        raise HTTPException(status_code=404, detail="Original transaction not found")
    
    if original_tx["status"] == "voided":
        raise HTTPException(status_code=400, detail="Cannot create credit note for voided transaction")
    
    # Validate items are from original transaction
    original_items = {item["product_id"]: item for item in original_tx["items"]}
    for item in data.items:
        if item.product_id not in original_items:
            raise HTTPException(status_code=400, detail=f"Product {item.product_id} not in original transaction")
        orig_item = original_items[item.product_id]
        if item.quantity > orig_item["quantity"]:
            raise HTTPException(status_code=400, detail=f"Credit quantity exceeds original quantity for {item.product_name}")
    
    # Calculate totals
    subtotal = sum(item.line_total for item in data.items)
    tax_amount = subtotal * (original_tx.get("tax_amount", 0) / original_tx.get("subtotal", 1)) if original_tx.get("subtotal") else 0
    total = subtotal + tax_amount
    
    credit_note_number = await generate_credit_note_number(user["organization_id"])
    
    credit_note = CreditNote(
        organization_id=user["organization_id"],
        store_id=store_id,
        credit_note_number=credit_note_number,
        original_transaction_id=data.original_transaction_id,
        original_receipt_number=original_tx["receipt_number"],
        items=[item.model_dump() for item in data.items],
        subtotal=subtotal,
        tax_amount=round(tax_amount, 2),
        total=round(total, 2),
        reason=data.reason,
        issued_by=user["id"],
        issued_by_name=f"{user['first_name']} {user['last_name']}"
    )
    
    await db.credit_notes.insert_one(serialize_datetime(credit_note.model_dump()))
    
    # Return items to stock
    for item in data.items:
        movement = StockMovementCreate(
            product_id=item.product_id,
            movement_type=StockMovementType.RETURN,
            quantity=item.quantity,
            reason=f"Credit Note {credit_note_number}"
        )
        await create_stock_movement(store_id, movement, user)
    
    return credit_note

@api_router.get("/credit-notes")
async def get_credit_notes(
    store_id: Optional[str] = None,
    status: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Get credit notes"""
    query = {"organization_id": user["organization_id"]}
    if store_id:
        query["store_id"] = store_id
    if status:
        query["status"] = status
    
    notes = await db.credit_notes.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [deserialize_datetime(n) for n in notes]

@api_router.get("/credit-notes/{credit_note_id}")
async def get_credit_note(
    credit_note_id: str,
    user: Dict = Depends(get_current_user)
):
    """Get a specific credit note"""
    note = await db.credit_notes.find_one({
        "id": credit_note_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not note:
        raise HTTPException(status_code=404, detail="Credit note not found")
    
    return deserialize_datetime(note)

@api_router.put("/credit-notes/{credit_note_id}/use")
async def use_credit_note(
    credit_note_id: str,
    transaction_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """Mark credit note as used in a transaction"""
    result = await db.credit_notes.update_one(
        {
            "id": credit_note_id,
            "organization_id": user["organization_id"],
            "status": "issued"
        },
        {"$set": {
            "status": "used",
            "used_transaction_id": transaction_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Credit note not found or already used")
    
    return {"message": "Credit note marked as used"}

# ==================== GOODS RECEIVED NOTE ROUTES ====================

async def generate_grn_number(org_id: str) -> str:
    """Generate sequential GRN number"""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.goods_received.count_documents({
        "organization_id": org_id,
        "grn_number": {"$regex": f"^GRN-{today}"}
    })
    return f"GRN-{today}-{str(count + 1).zfill(4)}"

@api_router.post("/goods-received")
async def record_goods_received(
    transfer_id: str = Query(...),
    notes: str = "",
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Record goods received from a warehouse transfer"""
    # Get the transfer
    transfer = await db.warehouse_transfers.find_one({
        "id": transfer_id,
        "organization_id": user["organization_id"]
    })
    
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    
    if transfer["status"] != "received":
        raise HTTPException(status_code=400, detail="Transfer must be in 'received' status")
    
    # Check if GRN already exists for this transfer
    existing = await db.goods_received.find_one({"transfer_id": transfer_id})
    if existing:
        raise HTTPException(status_code=400, detail="GRN already exists for this transfer")
    
    # Get warehouse info
    warehouse = await db.warehouses.find_one({"id": transfer["source_warehouse_id"]}, {"_id": 0})
    
    grn_number = await generate_grn_number(user["organization_id"])
    
    grn = GoodsReceivedNote(
        organization_id=user["organization_id"],
        store_id=transfer["destination_store_id"],
        grn_number=grn_number,
        transfer_id=transfer_id,
        transfer_number=transfer["transfer_number"],
        source_warehouse_id=transfer["source_warehouse_id"],
        source_warehouse_name=warehouse["name"] if warehouse else "Unknown",
        items=transfer["items"],
        received_by=user["id"],
        received_by_name=f"{user['first_name']} {user['last_name']}",
        notes=notes
    )
    
    await db.goods_received.insert_one(serialize_datetime(grn.model_dump()))
    return grn

@api_router.get("/goods-received")
async def get_goods_received(
    store_id: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Get goods received notes"""
    query = {"organization_id": user["organization_id"]}
    if store_id:
        query["store_id"] = store_id
    
    grns = await db.goods_received.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [deserialize_datetime(g) for g in grns]

@api_router.get("/goods-received/{grn_id}")
async def get_grn(
    grn_id: str,
    user: Dict = Depends(get_current_user)
):
    """Get a specific GRN"""
    grn = await db.goods_received.find_one({
        "id": grn_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")
    
    return deserialize_datetime(grn)

# ==================== STOCK AUDIT ROUTES ====================

@api_router.post("/audits", response_model=StockAudit)
async def create_audit(
    store_id: str = Query(...),
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Start a new stock audit"""
    audit_number = await generate_audit_number(user["organization_id"])
    
    # Get current stock for the store
    stock_items = await db.stock.aggregate([
        {"$match": {"store_id": store_id, "organization_id": user["organization_id"]}},
        {"$lookup": {
            "from": "products",
            "localField": "product_id",
            "foreignField": "id",
            "as": "product"
        }},
        {"$unwind": "$product"}
    ]).to_list(1000)
    
    items = [
        AuditItem(
            product_id=s["product_id"],
            product_name=s["product"]["name"],
            sku=s["product"]["sku"],
            system_quantity=s["quantity"],
            counted_quantity=0
        )
        for s in stock_items
    ]
    
    audit = StockAudit(
        organization_id=user["organization_id"],
        store_id=store_id,
        audit_number=audit_number,
        items=items,
        started_by=user["id"]
    )
    
    await db.stock_audits.insert_one(serialize_datetime(audit.model_dump()))
    return audit

@api_router.get("/audits", response_model=List[StockAudit])
async def get_audits(
    store_id: Optional[str] = None,
    status: Optional[str] = None,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get stock audits"""
    query = {"organization_id": user["organization_id"]}
    if store_id:
        query["store_id"] = store_id
    if status:
        query["status"] = status
    
    audits = await db.stock_audits.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return [deserialize_datetime(a) for a in audits]

@api_router.put("/audits/{audit_id}/count", response_model=StockAudit)
async def update_audit_count(
    audit_id: str,
    items: List[Dict[str, Any]],
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Update counted quantities"""
    audit = await db.stock_audits.find_one({
        "id": audit_id,
        "organization_id": user["organization_id"]
    })
    
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    
    # Update counted quantities
    for update_item in items:
        for audit_item in audit["items"]:
            if audit_item["product_id"] == update_item["product_id"]:
                audit_item["counted_quantity"] = update_item["counted_quantity"]
                audit_item["variance"] = update_item["counted_quantity"] - audit_item["system_quantity"]
    
    await db.stock_audits.update_one(
        {"id": audit_id},
        {"$set": {
            "items": audit["items"],
            "status": "in_progress",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.stock_audits.find_one({"id": audit_id}, {"_id": 0})
    return deserialize_datetime(updated)

@api_router.put("/audits/{audit_id}/complete", response_model=StockAudit)
async def complete_audit(
    audit_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Complete an audit"""
    await db.stock_audits.update_one(
        {"id": audit_id, "organization_id": user["organization_id"]},
        {"$set": {
            "status": "completed",
            "completed_by": user["id"],
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.stock_audits.find_one({"id": audit_id}, {"_id": 0})
    return deserialize_datetime(updated)

@api_router.put("/audits/{audit_id}/approve", response_model=StockAudit)
async def approve_audit(
    audit_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Approve an audit and adjust stock"""
    audit = await db.stock_audits.find_one({
        "id": audit_id,
        "organization_id": user["organization_id"]
    })
    
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    
    if audit["status"] != "completed":
        raise HTTPException(status_code=400, detail="Audit must be completed first")
    
    # Adjust stock based on variance
    for item in audit["items"]:
        if item["variance"] != 0:
            movement = StockMovementCreate(
                product_id=item["product_id"],
                movement_type=StockMovementType.ADJUSTMENT,
                quantity=item["counted_quantity"],
                reason=f"Audit adjustment {audit['audit_number']}"
            )
            await create_stock_movement(audit["store_id"], movement, user)
    
    await db.stock_audits.update_one(
        {"id": audit_id},
        {"$set": {
            "status": "approved",
            "approved_by": user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.stock_audits.find_one({"id": audit_id}, {"_id": 0})
    return deserialize_datetime(updated)

# ==================== ANALYTICS ROUTES ====================

@api_router.get("/analytics/sales-summary")
async def get_sales_summary(
    store_id: Optional[str] = None,
    period: str = "daily",  # daily, weekly, monthly, yearly
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get sales summary"""
    now = datetime.now(timezone.utc)
    
    # Calculate date range based on period
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=365)
    
    query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    if store_id:
        query["store_id"] = store_id
    
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$store_id",
            "gross_sales": {"$sum": "$total"},
            "tax_collected": {"$sum": "$tax_amount"},
            "transaction_count": {"$sum": 1},
            "average_transaction": {"$avg": "$total"}
        }}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(100)
    
    # Get store names
    store_ids = [r["_id"] for r in results]
    stores = await db.stores.find({"id": {"$in": store_ids}}, {"_id": 0}).to_list(100)
    store_map = {s["id"]: s["name"] for s in stores}
    
    return [{
        "store_id": r["_id"],
        "store_name": store_map.get(r["_id"], "Unknown"),
        "gross_sales": round(r["gross_sales"], 2),
        "tax_collected": round(r["tax_collected"], 2),
        "transaction_count": r["transaction_count"],
        "average_transaction": round(r["average_transaction"], 2)
    } for r in results]

@api_router.get("/analytics/stores-map")
async def get_stores_map_data(
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN]))
):
    """Get store data with sales for map visualization"""
    now = datetime.now(timezone.utc)
    
    # Get all stores with locations
    stores = await db.stores.find({
        "organization_id": user["organization_id"],
        "is_active": True
    }, {"_id": 0}).to_list(100)
    
    result = []
    for store in stores:
        # Get sales data for different periods
        daily_sales = await db.transactions.aggregate([
            {"$match": {
                "store_id": store["id"],
                "status": TransactionStatus.COMPLETED.value,
                "created_at": {"$gte": (now - timedelta(days=1)).isoformat()}
            }},
            {"$group": {"_id": None, "total": {"$sum": "$total"}}}
        ]).to_list(1)
        
        weekly_sales = await db.transactions.aggregate([
            {"$match": {
                "store_id": store["id"],
                "status": TransactionStatus.COMPLETED.value,
                "created_at": {"$gte": (now - timedelta(days=7)).isoformat()}
            }},
            {"$group": {"_id": None, "total": {"$sum": "$total"}}}
        ]).to_list(1)
        
        monthly_sales = await db.transactions.aggregate([
            {"$match": {
                "store_id": store["id"],
                "status": TransactionStatus.COMPLETED.value,
                "created_at": {"$gte": (now - timedelta(days=30)).isoformat()}
            }},
            {"$group": {"_id": None, "total": {"$sum": "$total"}}}
        ]).to_list(1)
        
        result.append({
            "store_id": store["id"],
            "store_name": store["name"],
            "location": store.get("location", {}),
            "daily_sales": daily_sales[0]["total"] if daily_sales else 0,
            "weekly_sales": weekly_sales[0]["total"] if weekly_sales else 0,
            "monthly_sales": monthly_sales[0]["total"] if monthly_sales else 0,
            "last_sync_at": store.get("last_sync_at")
        })
    
    return result

@api_router.get("/analytics/sales-trend")
async def get_sales_trend(
    store_id: Optional[str] = None,
    days: int = 30,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get daily sales trend"""
    now = datetime.now(timezone.utc)
    start_date = now - timedelta(days=days)
    
    query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    if store_id:
        query["store_id"] = store_id
    
    # Get all transactions in range
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    # Group by date
    daily_data = {}
    for t in transactions:
        date_str = t["created_at"][:10]  # Get YYYY-MM-DD
        if date_str not in daily_data:
            daily_data[date_str] = {"sales": 0, "tax": 0, "count": 0}
        daily_data[date_str]["sales"] += t["total"]
        daily_data[date_str]["tax"] += t["tax_amount"]
        daily_data[date_str]["count"] += 1
    
    # Convert to list sorted by date
    result = [
        {
            "date": date,
            "sales": round(data["sales"], 2),
            "tax": round(data["tax"], 2),
            "transaction_count": data["count"]
        }
        for date, data in sorted(daily_data.items())
    ]
    
    return result

@api_router.get("/analytics/top-products")
async def get_top_products(
    store_id: Optional[str] = None,
    limit: int = 10,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get top selling products"""
    now = datetime.now(timezone.utc)
    start_date = now - timedelta(days=30)
    
    match_query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    if store_id:
        # Verify store belongs to user's organization
        store = await db.stores.find_one({
            "id": store_id,
            "organization_id": user["organization_id"]
        }, {"id": 1})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found or access denied")
        match_query["store_id"] = store_id
    
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "product_name": {"$first": "$items.product_name"},
            "sku": {"$first": "$items.sku"},
            "total_quantity": {"$sum": "$items.quantity"},
            "total_revenue": {"$sum": "$items.line_total"}
        }},
        {"$sort": {"total_revenue": -1}},
        {"$limit": limit}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(limit)
    
    return [{
        "product_id": r["_id"],
        "product_name": r["product_name"],
        "sku": r["sku"],
        "total_quantity": r["total_quantity"],
        "total_revenue": round(r["total_revenue"], 2)
    } for r in results]

@api_router.get("/analytics/dashboard")
async def get_dashboard_data(
    store_id: Optional[str] = None,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get comprehensive dashboard data"""
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    base_query = {"organization_id": user["organization_id"]}
    if store_id:
        # Verify store belongs to user's organization
        store = await db.stores.find_one({
            "id": store_id,
            "organization_id": user["organization_id"]
        }, {"id": 1})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found or access denied")
        base_query["store_id"] = store_id
    
    # Today's sales
    today_query = {
        **base_query,
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": today_start.isoformat()}
    }
    today_sales = await db.transactions.aggregate([
        {"$match": today_query},
        {"$group": {
            "_id": None,
            "total": {"$sum": "$total"},
            "tax": {"$sum": "$tax_amount"},
            "count": {"$sum": 1}
        }}
    ]).to_list(1)
    
    # This week's sales
    week_start = now - timedelta(days=7)
    week_query = {
        **base_query,
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": week_start.isoformat()}
    }
    week_sales = await db.transactions.aggregate([
        {"$match": week_query},
        {"$group": {
            "_id": None,
            "total": {"$sum": "$total"},
            "count": {"$sum": 1}
        }}
    ]).to_list(1)
    
    # This month's sales
    month_start = now - timedelta(days=30)
    month_query = {
        **base_query,
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": month_start.isoformat()}
    }
    month_sales = await db.transactions.aggregate([
        {"$match": month_query},
        {"$group": {
            "_id": None,
            "total": {"$sum": "$total"},
            "count": {"$sum": 1}
        }}
    ]).to_list(1)
    
    # Low stock alerts
    low_stock_query = base_query.copy()
    if "store_id" not in low_stock_query and user["role"] != UserRole.SUPER_ADMIN:
        # Get all stores for org
        stores = await db.stores.find({"organization_id": user["organization_id"]}, {"id": 1}).to_list(100)
        low_stock_query["store_id"] = {"$in": [s["id"] for s in stores]}
    
    low_stock = await db.stock.aggregate([
        {"$match": low_stock_query},
        {"$match": {"$expr": {"$lte": ["$quantity", "$reorder_level"]}}},
        {"$lookup": {
            "from": "products",
            "localField": "product_id",
            "foreignField": "id",
            "as": "product"
        }},
        {"$unwind": "$product"},
        {"$project": {
            "_id": 0,
            "product_name": "$product.name",
            "sku": "$product.sku",
            "quantity": 1,
            "reorder_level": 1,
            "store_id": 1
        }},
        {"$limit": 10}
    ]).to_list(10)
    
    # Recent transactions
    recent_transactions = await db.transactions.find(
        {**base_query, "status": TransactionStatus.COMPLETED.value},
        {"_id": 0}
    ).sort("created_at", -1).to_list(5)
    
    # Store count
    store_count = await db.stores.count_documents({"organization_id": user["organization_id"], "is_active": True})
    
    # Product count
    product_count = await db.products.count_documents({"organization_id": user["organization_id"], "is_active": True})
    
    return {
        "today": {
            "sales": round(today_sales[0]["total"], 2) if today_sales else 0,
            "tax": round(today_sales[0]["tax"], 2) if today_sales else 0,
            "transactions": today_sales[0]["count"] if today_sales else 0
        },
        "week": {
            "sales": round(week_sales[0]["total"], 2) if week_sales else 0,
            "transactions": week_sales[0]["count"] if week_sales else 0
        },
        "month": {
            "sales": round(month_sales[0]["total"], 2) if month_sales else 0,
            "transactions": month_sales[0]["count"] if month_sales else 0
        },
        "low_stock_alerts": low_stock,
        "recent_transactions": [deserialize_datetime(t) for t in recent_transactions],
        "store_count": store_count,
        "product_count": product_count
    }

@api_router.get("/analytics/sales-per-product")
async def get_sales_per_product(
    store_id: Optional[str] = None,
    period: str = "monthly",  # daily, weekly, monthly, yearly
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get sales per product with filters"""
    now = datetime.now(timezone.utc)
    
    # Calculate date range based on period
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=365)
    
    match_query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    if store_id:
        # Verify store belongs to user's organization
        store = await db.stores.find_one({
            "id": store_id,
            "organization_id": user["organization_id"]
        }, {"id": 1})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found or access denied")
        match_query["store_id"] = store_id
    
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "product_name": {"$first": "$items.product_name"},
            "sku": {"$first": "$items.sku"},
            "brand": {"$first": "$items.brand"},
            "total_quantity": {"$sum": "$items.quantity"},
            "total_sales": {"$sum": "$items.line_total"},
            "total_cost": {"$sum": {"$multiply": ["$items.quantity", "$items.unit_price"]}}  # Will be updated with actual cost
        }},
        {"$sort": {"total_sales": -1}}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(1000)
    
    # Get product costs to calculate profit (only from user's organization)
    product_ids = [r["_id"] for r in results]
    products = await db.products.find({
        "id": {"$in": product_ids},
        "organization_id": user["organization_id"]
    }, {"_id": 0, "id": 1, "cost_price": 1}).to_list(1000)
    cost_map = {p["id"]: p.get("cost_price", 0) for p in products}
    
    return [{
        "product_id": r["_id"],
        "product_name": r["product_name"],
        "sku": r.get("sku", ""),
        "brand": r.get("brand", ""),
        "quantity_sold": r["total_quantity"],
        "total_sales": round(r["total_sales"], 2),
        "total_cost": round(r["total_quantity"] * cost_map.get(r["_id"], 0), 2),
        "profit": round(r["total_sales"] - (r["total_quantity"] * cost_map.get(r["_id"], 0)), 2)
    } for r in results]

@api_router.get("/analytics/profit-per-product")
async def get_profit_per_product(
    store_id: Optional[str] = None,
    period: str = "monthly",
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get profit per product with filters (same as sales-per-product but sorted by profit)"""
    now = datetime.now(timezone.utc)
    
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=365)
    
    match_query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    if store_id:
        # Verify store belongs to user's organization
        store = await db.stores.find_one({
            "id": store_id,
            "organization_id": user["organization_id"]
        }, {"id": 1})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found or access denied")
        match_query["store_id"] = store_id
    
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "product_name": {"$first": "$items.product_name"},
            "sku": {"$first": "$items.sku"},
            "brand": {"$first": "$items.brand"},
            "total_quantity": {"$sum": "$items.quantity"},
            "total_sales": {"$sum": "$items.line_total"}
        }},
        {"$sort": {"total_sales": -1}}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(1000)
    
    # Get product costs (only from user's organization)
    product_ids = [r["_id"] for r in results]
    products = await db.products.find({
        "id": {"$in": product_ids},
        "organization_id": user["organization_id"]
    }, {"_id": 0, "id": 1, "cost_price": 1}).to_list(1000)
    cost_map = {p["id"]: p.get("cost_price", 0) for p in products}
    
    # Calculate profit for each product
    products_with_profit = []
    for r in results:
        cost = cost_map.get(r["_id"], 0)
        total_cost = r["total_quantity"] * cost
        profit = r["total_sales"] - total_cost
        products_with_profit.append({
            "product_id": r["_id"],
            "product_name": r["product_name"],
            "sku": r.get("sku", ""),
            "brand": r.get("brand", ""),
            "quantity_sold": r["total_quantity"],
            "total_sales": round(r["total_sales"], 2),
            "total_cost": round(total_cost, 2),
            "profit": round(profit, 2)
        })
    
    # Sort by profit descending
    products_with_profit.sort(key=lambda x: x["profit"], reverse=True)
    
    return products_with_profit

@api_router.get("/analytics/sales-per-branch")
async def get_sales_per_branch(
    period: str = "monthly",
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get sales per branch/store with period filter"""
    now = datetime.now(timezone.utc)
    
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=365)
    
    match_query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$store_id",
            "total_sales": {"$sum": "$total"},
            "total_tax": {"$sum": "$tax_amount"},
            "transaction_count": {"$sum": 1}
        }},
        {"$sort": {"total_sales": -1}}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(100)
    
    # Get store names (only from user's organization)
    store_ids = [r["_id"] for r in results]
    stores = await db.stores.find({
        "id": {"$in": store_ids},
        "organization_id": user["organization_id"]
    }, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    store_map = {s["id"]: s["name"] for s in stores}
    
    return [{
        "store_id": r["_id"],
        "store_name": store_map.get(r["_id"], "Unknown"),
        "total_sales": round(r["total_sales"], 2),
        "total_tax": round(r["total_tax"], 2),
        "transaction_count": r["transaction_count"]
    } for r in results]

@api_router.get("/analytics/profit-per-branch")
async def get_profit_per_branch(
    period: str = "monthly",
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN, UserRole.ORG_ADMIN, UserRole.STORE_ADMIN]))
):
    """Get profit per branch/store with period filter"""
    now = datetime.now(timezone.utc)
    
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=365)
    
    match_query = {
        "organization_id": user["organization_id"],
        "status": TransactionStatus.COMPLETED.value,
        "created_at": {"$gte": start_date.isoformat()}
    }
    
    # Get transactions with items
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$store_id",
            "total_sales": {"$sum": "$items.line_total"},
            "items": {"$push": {
                "product_id": "$items.product_id",
                "quantity": "$items.quantity"
            }}
        }}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(100)
    
    # Get store names (only from user's organization)
    store_ids = [r["_id"] for r in results]
    stores = await db.stores.find({
        "id": {"$in": store_ids},
        "organization_id": user["organization_id"]
    }, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    store_map = {s["id"]: s["name"] for s in stores}
    
    # Get product costs (only from user's organization)
    all_product_ids = set()
    for r in results:
        for item in r["items"]:
            all_product_ids.add(item["product_id"])
    
    products = await db.products.find({
        "id": {"$in": list(all_product_ids)},
        "organization_id": user["organization_id"]
    }, {"_id": 0, "id": 1, "cost_price": 1}).to_list(1000)
    cost_map = {p["id"]: p.get("cost_price", 0) for p in products}
    
    # Calculate profit for each branch
    branches_with_profit = []
    for r in results:
        total_cost = 0
        for item in r["items"]:
            cost = cost_map.get(item["product_id"], 0)
            total_cost += item["quantity"] * cost
        
        profit = r["total_sales"] - total_cost
        branches_with_profit.append({
            "store_id": r["_id"],
            "store_name": store_map.get(r["_id"], "Unknown"),
            "total_sales": round(r["total_sales"], 2),
            "total_cost": round(total_cost, 2),
            "profit": round(profit, 2)
        })
    
    # Sort by profit descending
    branches_with_profit.sort(key=lambda x: x["profit"], reverse=True)
    
    return branches_with_profit

# ==================== SYNC ROUTES ====================

@api_router.post("/sync/push")
async def push_sync(
    data: Dict[str, Any],
    store_id: str = Query(...),
    user: Dict = Depends(get_current_user)
):
    """Push offline data to server"""
    sync_log = SyncLog(
        organization_id=user["organization_id"],
        store_id=store_id,
        sync_type="push",
        entity_type=data.get("entity_type", "mixed"),
        status="in_progress"
    )
    
    try:
        records_synced = 0
        
        # Process transactions
        if "transactions" in data:
            for transaction in data["transactions"]:
                existing = await db.transactions.find_one({"local_id": transaction.get("local_id")})
                if not existing:
                    transaction["synced"] = True
                    await db.transactions.insert_one(serialize_datetime(transaction))
                    records_synced += 1
        
        # Process stock movements
        if "stock_movements" in data:
            for movement in data["stock_movements"]:
                await db.stock_movements.insert_one(serialize_datetime(movement))
                records_synced += 1
        
        sync_log.records_synced = records_synced
        sync_log.status = "completed"
        
        # Update store last sync
        await db.stores.update_one(
            {"id": store_id},
            {"$set": {"last_sync_at": datetime.now(timezone.utc).isoformat()}}
        )
        
    except Exception as e:
        sync_log.status = "failed"
        sync_log.error_message = str(e)
    
    await db.sync_logs.insert_one(serialize_datetime(sync_log.model_dump()))
    
    return {"status": sync_log.status, "records_synced": sync_log.records_synced}

@api_router.get("/sync/pull")
async def pull_sync(
    store_id: str = Query(...),
    last_sync: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Pull data from server for offline cache"""
    query = {"organization_id": user["organization_id"]}
    
    if last_sync:
        query["updated_at"] = {"$gt": last_sync}
    
    # Get products
    products = await db.products.find(query, {"_id": 0}).to_list(10000)
    
    # Get stock for this store
    stock_query = {**query, "store_id": store_id}
    stock = await db.stock.find(stock_query, {"_id": 0}).to_list(10000)
    
    # Get organization settings
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    
    return {
        "products": products,
        "stock": stock,
        "organization": org,
        "sync_timestamp": datetime.now(timezone.utc).isoformat()
    }

# ==================== PRINT ROUTES ====================

@api_router.post("/print/receipt")
async def print_receipt(
    transaction_id: str,
    printer_id: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Generate receipt data for printing"""
    transaction = await db.transactions.find_one({
        "id": transaction_id,
        "organization_id": user["organization_id"]
    }, {"_id": 0})
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    store = await db.stores.find_one({"id": transaction["store_id"]}, {"_id": 0})
    
    # Generate ESC/POS compatible receipt data
    receipt_lines = []
    currency = org["settings"].get("currency_symbol", "K")
    
    # Header
    receipt_lines.append({"type": "text", "align": "center", "bold": True, "size": 2, "text": org["name"]})
    if org["settings"].get("tpin"):
        receipt_lines.append({"type": "text", "align": "center", "text": f"TPIN: {org['settings']['tpin']}"})
    receipt_lines.append({"type": "text", "align": "center", "text": store["name"]})
    if store.get("address", {}).get("street"):
        receipt_lines.append({"type": "text", "align": "center", "text": store["address"]["street"]})
    if store.get("phone"):
        receipt_lines.append({"type": "text", "align": "center", "text": f"Tel: {store['phone']}"})
    
    receipt_lines.append({"type": "separator"})
    
    # Transaction info
    receipt_lines.append({"type": "text", "text": f"Receipt #: {transaction['receipt_number']}"})
    receipt_lines.append({"type": "text", "text": f"Date: {transaction['created_at'][:19].replace('T', ' ')}"})
    receipt_lines.append({"type": "text", "text": f"Cashier: {transaction['cashier_name']}"})
    if transaction.get("customer_name"):
        receipt_lines.append({"type": "text", "text": f"Customer: {transaction['customer_name']}"})
    
    receipt_lines.append({"type": "separator"})
    
    # Items
    for item in transaction["items"]:
        receipt_lines.append({
            "type": "item",
            "name": item["product_name"],
            "qty": item["quantity"],
            "price": item["unit_price"],
            "total": item["line_total"],
            "currency": currency
        })
    
    receipt_lines.append({"type": "separator"})
    
    # Totals
    receipt_lines.append({"type": "total", "label": "Subtotal", "value": transaction["subtotal"], "currency": currency})
    if transaction.get("discount_amount", 0) > 0:
        receipt_lines.append({"type": "total", "label": "Discount", "value": -transaction["discount_amount"], "currency": currency})
    receipt_lines.append({"type": "total", "label": f"VAT ({org['settings'].get('tax_rate', 16)}%)", "value": transaction["tax_amount"], "currency": currency})
    receipt_lines.append({"type": "total", "label": "TOTAL", "value": transaction["total"], "currency": currency, "bold": True, "size": 2})
    
    receipt_lines.append({"type": "separator"})
    
    # Payment
    for payment in transaction.get("payments", []):
        receipt_lines.append({"type": "text", "text": f"Paid ({payment['method'].upper()}): {currency}{payment['amount']:.2f}"})
    
    receipt_lines.append({"type": "separator"})
    
    # Footer
    receipt_lines.append({"type": "text", "align": "center", "text": org["settings"].get("receipt_footer", "Thank you!")})
    receipt_lines.append({"type": "text", "align": "center", "text": f"Printed: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}"})
    
    # Barcode/QR for receipt number
    receipt_lines.append({"type": "barcode", "data": transaction["receipt_number"]})
    
    return {
        "receipt_id": transaction["id"],
        "receipt_number": transaction["receipt_number"],
        "lines": receipt_lines,
        "raw_data": {
            "header": {
                "business_name": org["name"],
                "tpin": org["settings"].get("tpin", ""),
                "store_name": store["name"],
                "address": store.get("address", {}),
                "phone": store.get("phone", "")
            },
            "transaction": transaction,
            "footer": {
                "message": org["settings"].get("receipt_footer", "Thank you for your business!"),
                "currency": org["settings"].get("currency", "ZMW")
            }
        },
        "printer_config": {
            "paper_width": 80,
            "charset": "UTF-8",
            "cut_paper": True
        }
    }

@api_router.get("/printers")
async def get_printers(user: Dict = Depends(get_current_user)):
    """Get configured printers for organization"""
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    configured_printers = org.get("settings", {}).get("printers", []) if org else []
    
    return {
        "configured_printers": configured_printers,
        "available_printers": [
            {"id": "usb-default", "name": "USB Thermal Printer", "type": "usb", "status": "available"},
            {"id": "bt-default", "name": "Bluetooth Thermal Printer", "type": "bluetooth", "status": "available"},
            {"id": "network-default", "name": "Network Printer", "type": "network", "status": "available"}
        ],
        "note": "Connect actual printers via USB, Bluetooth, or Network for real printing."
    }

@api_router.post("/printers/test")
async def test_printer(
    printer_id: str,
    user: Dict = Depends(get_current_user)
):
    """Test printer connection"""
    # This would connect to actual printer in production
    return {
        "status": "success",
        "message": f"Test page sent to printer {printer_id}",
        "note": "In production, this would send actual ESC/POS commands to the printer"
    }

# ==================== SAAS ADMIN ROUTES ====================

@api_router.post("/admin/super-admin", response_model=UserResponse)
async def create_super_admin(
    data: UserCreate,
    admin_key: str = Query(...)
):
    """Create a super admin user (requires admin key)"""
    # Simple admin key check - in production, use environment variable
    if admin_key != os.environ.get("SUPER_ADMIN_KEY", "stockmaster-admin-2024"):
        raise HTTPException(status_code=403, detail="Invalid admin key")
    
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        email=data.email,
        password_hash=hash_password(data.password),
        first_name=data.first_name,
        last_name=data.last_name,
        role=UserRole.SUPER_ADMIN,
        organization_id=None
    )
    
    await db.users.insert_one(serialize_datetime(user.model_dump()))
    return UserResponse(**user.model_dump())

@api_router.get("/admin/organizations")
async def admin_get_all_organizations(
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Get all organizations with stats (Super Admin only)"""
    orgs = await db.organizations.find({}, {"_id": 0}).to_list(1000)
    
    result = []
    for org in orgs:
        # Get stats for each org
        store_count = await db.stores.count_documents({"organization_id": org["id"]})
        user_count = await db.users.count_documents({"organization_id": org["id"]})
        product_count = await db.products.count_documents({"organization_id": org["id"]})
        
        # Get total sales (last 30 days)
        thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
        sales_agg = await db.transactions.aggregate([
            {"$match": {
                "organization_id": org["id"],
                "status": "completed",
                "created_at": {"$gte": thirty_days_ago}
            }},
            {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
        ]).to_list(1)
        
        result.append({
            **deserialize_datetime(org),
            "stats": {
                "store_count": store_count,
                "user_count": user_count,
                "product_count": product_count,
                "monthly_sales": sales_agg[0]["total"] if sales_agg else 0,
                "monthly_transactions": sales_agg[0]["count"] if sales_agg else 0
            }
        })
    
    return result

@api_router.post("/admin/organizations")
async def admin_create_organization(
    data: OrganizationCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Create a new organization (Super Admin only)"""
    existing = await db.organizations.find_one({"slug": data.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Organization slug already exists")
    
    # Filter out None values to let Organization model use defaults
    org_data = {k: v for k, v in data.model_dump().items() if v is not None}
    org = Organization(**org_data)
    await db.organizations.insert_one(serialize_datetime(org.model_dump()))
    return org

@api_router.put("/admin/organizations/{org_id}")
async def admin_update_organization(
    org_id: str,
    data: Dict[str, Any],
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Update any organization (Super Admin only)"""
    if "settings" in data and isinstance(data.get("settings"), dict):
        existing = await db.organizations.find_one({"id": org_id}, {"_id": 0, "settings": 1})
        if existing:
            data["settings"] = {
                **existing.get("settings", {}),
                **data.get("settings", {})
            }
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.organizations.update_one(
        {"id": org_id},
        {"$set": serialize_datetime(data)}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    return deserialize_datetime(org)

@api_router.delete("/admin/organizations/{org_id}")
async def admin_delete_organization(
    org_id: str,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Deactivate an organization (Super Admin only)"""
    result = await db.organizations.update_one(
        {"id": org_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    return {"message": "Organization deactivated"}

@api_router.get("/admin/users")
async def admin_get_all_users(
    org_id: Optional[str] = None,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Get all users across organizations (Super Admin only)"""
    query = {}
    if org_id:
        query["organization_id"] = org_id
    
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [deserialize_datetime(u) for u in users]

@api_router.post("/admin/organizations/{org_id}/users")
async def admin_create_user_for_org(
    org_id: str,
    data: UserCreate,
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Create a user for specific organization (Super Admin only)"""
    org = await db.organizations.find_one({"id": org_id})
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    new_user = User(
        organization_id=org_id,
        email=data.email,
        password_hash=hash_password(data.password),
        first_name=data.first_name,
        last_name=data.last_name,
        role=data.role,
        store_ids=data.store_ids,
        pin=data.pin
    )
    
    await db.users.insert_one(serialize_datetime(new_user.model_dump()))
    return UserResponse(**new_user.model_dump())

@api_router.get("/admin/stats")
async def admin_get_platform_stats(
    user: Dict = Depends(require_role([UserRole.SUPER_ADMIN]))
):
    """Get overall platform statistics (Super Admin only)"""
    org_count = await db.organizations.count_documents({"is_active": True})
    total_orgs = await db.organizations.count_documents({})
    store_count = await db.stores.count_documents({"is_active": True})
    user_count = await db.users.count_documents({"is_active": True})
    product_count = await db.products.count_documents({"is_active": True})
    
    # Get total sales (last 30 days)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    sales_agg = await db.transactions.aggregate([
        {"$match": {"status": "completed", "created_at": {"$gte": thirty_days_ago}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    
    # Get daily signups for last 7 days
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    daily_signups = await db.users.aggregate([
        {"$match": {"created_at": {"$gte": seven_days_ago}}},
        {"$group": {
            "_id": {"$substr": ["$created_at", 0, 10]},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]).to_list(7)
    
    return {
        "organizations": {
            "active": org_count,
            "total": total_orgs
        },
        "stores": store_count,
        "users": user_count,
        "products": product_count,
        "monthly_sales": {
            "total": sales_agg[0]["total"] if sales_agg else 0,
            "transactions": sales_agg[0]["count"] if sales_agg else 0
        },
        "daily_signups": daily_signups
    }

# ==================== HEALTH CHECK ====================

@api_router.get("/health")
async def health_check():
    """Health check endpoint"""
    db_status = "unknown"
    try:
        # Check MongoDB connection
        await client.admin.command('ping')
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)[:50]}"
    
    return {
        "status": "healthy",
        "database": db_status,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": "1.0.0"
    }

# Include router and setup middleware
app.include_router(api_router)

# Add explicit OPTIONS handler for CORS preflight
@app.options("/{full_path:path}")
async def options_handler(full_path: str):
    """Handle CORS preflight requests"""
    return Response(
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, PATCH",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Allow-Credentials": "true",
        }
    )

# CORS middleware - must be added after router but before static file serving
cors_origins = os.environ.get('CORS_ORIGINS', '*')
if cors_origins == '*':
    # Allow all origins
    allow_origins = ['*']
else:
    # Split comma-separated origins
    allow_origins = [origin.strip() for origin in cors_origins.split(',') if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=allow_origins,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# Serve static frontend files (for Railway/production deployment)
STATIC_DIR = ROOT_DIR / "static"
if STATIC_DIR.exists():
    # Mount static assets (JS, CSS, images, etc.) from the static subdirectory
    static_assets_dir = STATIC_DIR / "static"
    if static_assets_dir.exists():
        app.mount("/static", StaticFiles(directory=str(static_assets_dir)), name="static_assets")
    
    # Serve index.html for all non-API routes (SPA support)
    # Only handle GET requests - POST/PUT/DELETE go to API router
    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        """Serve the React SPA for all non-API routes"""
        # Don't intercept API routes (shouldn't happen with GET, but safety check)
        if full_path.startswith("api/"):
            logger.warning(f"GET request to API route: /{full_path} - returning 404")
            raise HTTPException(status_code=404, detail="Not found")
        
        # Don't intercept /static routes (handled by mount above)
        if full_path.startswith("static/"):
            raise HTTPException(status_code=404, detail="Not found")
        
        # Serve static files if they exist (like manifest.json, favicon.ico, etc.)
        file_path = STATIC_DIR / full_path
        if file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        
        # Fallback to index.html for SPA routing
        index_path = STATIC_DIR / "index.html"
        if index_path.exists():
            return FileResponse(str(index_path))
        
        raise HTTPException(status_code=404, detail="Not found")

@app.on_event("startup")
async def startup_db_client():
    """Create indexes on startup"""
    logger.info("=" * 60)
    logger.info("EmergePOS Server Starting...")
    logger.info(f"Port: {os.environ.get('PORT', '8000')}")
    logger.info(f"Database: {os.environ.get('DB_NAME', 'pos_system')}")
    logger.info(f"MONGO_URL configured: {'Yes' if os.environ.get('MONGO_URL') else 'NO - MISSING!'}")
    logger.info("=" * 60)
    
    try:
        # Test MongoDB connection
        await client.admin.command('ping')
        logger.info("✓ MongoDB connection successful")
    except Exception as e:
        logger.error(f"✗ MongoDB connection failed: {e}")
        logger.error("The server will start but database operations will fail")
    
    try:
        # Users indexes
        await db.users.create_index("email", unique=True)
        await db.users.create_index("organization_id")
        
        # Organizations indexes
        await db.organizations.create_index("slug", unique=True)
        
        # Stores indexes
        await db.stores.create_index([("organization_id", 1), ("code", 1)], unique=True)
        
        # Products indexes
        await db.products.create_index([("organization_id", 1), ("sku", 1)], unique=True)
        await db.products.create_index([("organization_id", 1), ("barcode", 1)])
        
        # Stock indexes
        await db.stock.create_index([("organization_id", 1), ("store_id", 1), ("product_id", 1)], unique=True)
        
        # Transactions indexes
        await db.transactions.create_index([("organization_id", 1), ("store_id", 1), ("created_at", -1)])
        await db.transactions.create_index("local_id")
        await db.transactions.create_index("receipt_number")
        
        # Warehouse indexes
        await db.warehouses.create_index([("organization_id", 1), ("code", 1)], unique=True)
        await db.warehouse_stock.create_index([("organization_id", 1), ("warehouse_id", 1), ("product_id", 1)], unique=True)
        await db.warehouse_transfers.create_index([("organization_id", 1), ("created_at", -1)])
        
        logger.info("Database indexes created successfully")
        
        # Seed default users if none exist
        await seed_default_data()
    except Exception as e:
        logger.error(f"Error creating indexes: {e}")

async def seed_default_data():
    """Seed default users and organization if database is empty"""
    try:
        # Check if any users exist
        user_count = await db.users.count_documents({})
        if user_count > 0:
            logger.info("Database already has users, skipping seed")
            return
        
        logger.info("Seeding default data...")
        
        # Create organization
        org_id = str(uuid.uuid4())
        org = {
            "id": org_id,
            "name": "Demo Organization",
            "slug": "demo-org",
            "settings": {
                "currency": "ZMW",
                "currency_symbol": "K",
                "tax_rate": 16.0,
                "tax_inclusive_pricing": True,
                "receipt_footer": "Thank you for your business!",
                "invoice_prefix": "",
                "timezone": "Africa/Lusaka",
                "date_format": "DD/MM/YYYY",
                "fiscal_compliance_enabled": True,
                "payment_methods": [
                    {"id": str(uuid.uuid4()), "name": "Cash", "code": "cash", "icon": "💵", "is_active": True, "requires_reference": False},
                    {"id": str(uuid.uuid4()), "name": "Card", "code": "card", "icon": "💳", "is_active": True, "requires_reference": True},
                    {"id": str(uuid.uuid4()), "name": "Mobile Money", "code": "mobile_money", "icon": "📱", "is_active": True, "requires_reference": True},
                ],
                "printers": [],
                "allow_negative_stock": False,
                "low_stock_threshold": 10
            },
            "is_active": True,
            "subscription_plan": "premium",
            "max_stores": 10,
            "max_users": 50,
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }
        await db.organizations.insert_one(org)
        logger.info(f"Created organization: {org['name']}")
        
        # Create store
        store_id = str(uuid.uuid4())
        store = {
            "id": store_id,
            "organization_id": org_id,
            "name": "Main Store",
            "code": "MAIN001",
            "address": "123 Main Street, Lusaka",
            "phone": "+260-97-123-4567",
            "email": "mainstore@posystem.com",
            "is_active": True,
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }
        await db.stores.insert_one(store)
        logger.info(f"Created store: {store['name']}")
        
        # Create default users
        users = [
            {
                "id": str(uuid.uuid4()),
                "email": "superadmin@posystem.com",
                "password_hash": hash_password("SuperAdmin123!"),
                "first_name": "Super",
                "last_name": "Admin",
                "role": "super_admin",
                "organization_id": None,
                "store_ids": [],
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            },
            {
                "id": str(uuid.uuid4()),
                "email": "orgadmin@posystem.com",
                "password_hash": hash_password("OrgAdmin123!"),
                "first_name": "Org",
                "last_name": "Admin",
                "role": "org_admin",
                "organization_id": org_id,
                "store_ids": [],
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            },
            {
                "id": str(uuid.uuid4()),
                "email": "storeadmin@posystem.com",
                "password_hash": hash_password("StoreAdmin123!"),
                "first_name": "Store",
                "last_name": "Admin",
                "role": "store_admin",
                "organization_id": org_id,
                "store_ids": [store_id],
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            },
            {
                "id": str(uuid.uuid4()),
                "email": "cashier@posystem.com",
                "password_hash": hash_password("Cashier123!"),
                "first_name": "Test",
                "last_name": "Cashier",
                "role": "cashier",
                "organization_id": org_id,
                "store_ids": [store_id],
                "pin": "1234",
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
        ]
        
        await db.users.insert_many(users)
        logger.info(f"Created {len(users)} default users")
        
        # Create sample product
        product = {
            "id": str(uuid.uuid4()),
            "organization_id": org_id,
            "name": "Coca Cola 500ml",
            "sku": "COKE-500ML",
            "barcode": "1234567890123",
            "category": "Beverages",
            "brand": "Coca Cola",
            "cost_price": 2.50,
            "selling_price": 5.00,
            "tax_type": "standard",
            "unit": "piece",
            "is_active": True,
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }
        await db.products.insert_one(product)
        
        # Create stock for the product
        stock = {
            "id": str(uuid.uuid4()),
            "organization_id": org_id,
            "store_id": store_id,
            "product_id": product["id"],
            "quantity": 100,
            "reorder_level": 10,
            "last_updated": datetime.now(timezone.utc)
        }
        await db.stock.insert_one(stock)
        
        logger.info("Default data seeded successfully!")
        logger.info("Login credentials:")
        logger.info("  Super Admin: superadmin@posystem.com / SuperAdmin123!")
        logger.info("  Org Admin: orgadmin@posystem.com / OrgAdmin123!")
        logger.info("  Store Admin: storeadmin@posystem.com / StoreAdmin123!")
        logger.info("  Cashier: cashier@posystem.com / Cashier123!")
        
    except Exception as e:
        logger.error(f"Error seeding default data: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
